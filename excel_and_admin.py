# ==================== فایل گزارش اکسل و دستورات ادمین ====================
# این فایل شامل: تولید فایل اکسل، دستورات ادمین (/getexcel, /getdata, /summary,
# /broadcast, /reply) است
#
# نکته مهم: علاوه بر شیت‌های قبلی، این نسخه شامل موارد زیر است:
#  - شیت «اطلاعات کاربران» کامل شد: حیطه فعالیت، آیدی شبکه اجتماعی، یوزرنیم
#    تلگرام، تاریخ اولین بازدید و وضعیت/تاریخ انقضای اشتراک هم اضافه شدند.
#  - شیت جدید «سفارش محصولات» برای سفارش‌های آزاد از تب محصولات/خدمات.
#  - شیت جدید «وضعیت اشتراک» برای تاریخچه‌ی کامل اشتراک هر کاربر.
#  - دستور ادمین جدید /reply <آیدی کاربر> <متن پیام> برای پاسخ مستقیم به کاربر.

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from telegram import Update
from config import (
    read_json, USERS_FILE, SURVEY_FILE, ASSESSMENT_FILE, SECTION_FORMS_FILE,
    SUBSCRIPTIONS_FILE, PRODUCT_ORDERS_FILE, SUBSCRIPTION_TIERS, BANNED_USERS_FILE,
    EXCEL_FILE, ADMIN_ID, logger
)

# ==================== استایل مشترک هدر شیت‌ها ====================
def _style_header(ws, headers):
    """رنگ، فونت و عرض ستون‌های هدر یک شیت را تنظیم می‌کند"""
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

# ==================== تولید فایل اکسل ====================
def generate_excel_report():
    """ساخت فایل اکسل کامل: کاربران، فرم ارزیابی، پرسشنامه، فرم‌های بخش‌ها، سفارش‌ها، اشتراک و آمار"""
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    assessments = read_json(ASSESSMENT_FILE, {})
    section_forms = read_json(SECTION_FORMS_FILE, {})
    subscriptions = read_json(SUBSCRIPTIONS_FILE, {})
    product_orders = read_json(PRODUCT_ORDERS_FILE, {})
    banned_users = read_json(BANNED_USERS_FILE, [])

    wb = openpyxl.Workbook()

    # ===== شیت ۱: اطلاعات کاربران (کامل) =====
    ws1 = wb.active
    ws1.title = "اطلاعات کاربران"

    headers1 = ["ردیف", "آیدی تلگرام", "یوزرنیم تلگرام", "نام", "نام خانوادگی",
                "سن / تاریخ تولد", "شماره تماس", "شهر", "نام کسب و کار",
                "حیطه فعالیت", "آیدی شبکه اجتماعی / آدرس", "وضعیت اشتراک",
                "اعتبار اشتراک تا", "مسدود؟", "اولین بازدید", "آخرین بروزرسانی"]
    _style_header(ws1, headers1)

    tier_labels = {k: v["label"] for k, v in SUBSCRIPTION_TIERS.items()}
    status_labels_fa = {
        "active": "🟢 فعال", "pending": "🟡 در انتظار تایید",
        "expired": "🔴 منقضی", "rejected": "🔴 رد شده", "none": "⚪ ندارد",
    }

    for row, (user_id, info) in enumerate(users.items(), 2):
        sub = subscriptions.get(user_id, {})
        sub_status = sub.get("status", "none")
        sub_tier = tier_labels.get(sub.get("tier"), "-") if sub.get("tier") else "-"
        status_display = status_labels_fa.get(sub_status, sub_status)
        status_display = f"{status_display} ({sub_tier})" if sub.get("tier") else status_display
        is_banned = "🚫 بله" if user_id in banned_users else "خیر"

        ws1.cell(row=row, column=1, value=row - 1)
        ws1.cell(row=row, column=2, value=info.get('telegram_id', user_id))
        ws1.cell(row=row, column=3, value=info.get('telegram_username', ''))
        ws1.cell(row=row, column=4, value=info.get('first_name', ''))
        ws1.cell(row=row, column=5, value=info.get('last_name', ''))
        ws1.cell(row=row, column=6, value=info.get('birth_date', ''))
        ws1.cell(row=row, column=7, value=info.get('phone', ''))
        ws1.cell(row=row, column=8, value=info.get('city', ''))
        ws1.cell(row=row, column=9, value=info.get('business_name', ''))
        ws1.cell(row=row, column=10, value=info.get('field', ''))
        ws1.cell(row=row, column=11, value=info.get('social', ''))
        ws1.cell(row=row, column=12, value=status_display)
        ws1.cell(row=row, column=13, value=sub.get('expires_at', '-'))
        ws1.cell(row=row, column=14, value=is_banned)
        ws1.cell(row=row, column=15, value=info.get('first_seen', ''))
        ws1.cell(row=row, column=16, value=info.get('last_update', ''))

    # ===== شیت ۲: فرم ارزیابی بازار کار =====
    ws2 = wb.create_sheet("فرم_ارزیابی")

    headers2 = ["ردیف", "آیدی تلگرام", "نام و نام خانوادگی", "شماره تماس", "سن",
                "استان و شهر", "نقش فعلی", "حوزه تخصصی", "سه توانمندی",
                "دغدغه شغلی", "درآمد ماهیانه", "هدف یک ساله", "موانع اصلی",
                "مسئله حل شدنی", "نیاز اصلی", "نکته تکمیلی", "تاریخ ثبت"]
    _style_header(ws2, headers2)

    for row, (user_id, answers) in enumerate(assessments.items(), 2):
        user_info = users.get(user_id, {})
        ws2.cell(row=row, column=1, value=row - 1)
        ws2.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws2.cell(row=row, column=3, value=answers.get('full_name', ''))
        ws2.cell(row=row, column=4, value=answers.get('phone', ''))
        ws2.cell(row=row, column=5, value=answers.get('age', ''))
        ws2.cell(row=row, column=6, value=answers.get('location', ''))
        ws2.cell(row=row, column=7, value=answers.get('role', ''))
        ws2.cell(row=row, column=8, value=answers.get('field', ''))
        ws2.cell(row=row, column=9, value=answers.get('strengths', ''))
        ws2.cell(row=row, column=10, value=answers.get('challenge', ''))
        ws2.cell(row=row, column=11, value=answers.get('income', ''))
        ws2.cell(row=row, column=12, value=answers.get('goal', ''))
        ws2.cell(row=row, column=13, value=answers.get('obstacles', ''))
        ws2.cell(row=row, column=14, value=answers.get('solve_problem', ''))
        ws2.cell(row=row, column=15, value=answers.get('need', ''))
        ws2.cell(row=row, column=16, value=answers.get('note', ''))
        ws2.cell(row=row, column=17, value=answers.get('submitted_at', ''))

    # ===== شیت ۳: پرسشنامه تخصصی =====
    ws3 = wb.create_sheet("پرسشنامه")

    headers3 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "نام کسب و کار",
                "درباره کسب و کار", "محصولات و مزیت", "زیرساخت مجازی",
                "تیم", "فروش ماهیانه", "چالش اصلی", "نیاز مشاوره"]
    _style_header(ws3, headers3)

    for row, (user_id, answers) in enumerate(surveys.items(), 2):
        user_info = users.get(user_id, {})
        ws3.cell(row=row, column=1, value=row - 1)
        ws3.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws3.cell(row=row, column=3, value=user_info.get('first_name', ''))
        ws3.cell(row=row, column=4, value=user_info.get('last_name', ''))
        ws3.cell(row=row, column=5, value=user_info.get('business_name', ''))
        ws3.cell(row=row, column=6, value=answers.get('about_business', ''))
        ws3.cell(row=row, column=7, value=answers.get('products', ''))
        ws3.cell(row=row, column=8, value=answers.get('infrastructure', ''))
        ws3.cell(row=row, column=9, value=answers.get('team', ''))
        ws3.cell(row=row, column=10, value=answers.get('sales', ''))
        ws3.cell(row=row, column=11, value=answers.get('problem', ''))
        ws3.cell(row=row, column=12, value=answers.get('consulting', ''))

    # ===== شیت ۴: فرم‌های بخش‌ها (کسب‌وکار / مسئولیت اجتماعی / مسیر رشد) =====
    # هر ردیف یعنی یک فرم تکمیل‌شده توسط یک کاربر. چون هر کاربر می‌تواند
    # چند زیربخش مختلف را پر کند (مثلاً هم برند شخصی و هم توسعه شغلی)،
    # برای هرکدام یک ردیف جدا نوشته می‌شود. پاسخ هر سوال هم در یک ستون
    # مجزا (سوال ۱ تا ۵) قرار می‌گیرد تا خوانا باشد.
    ws4 = wb.create_sheet("فرم‌های بخش‌ها")

    headers4 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "نوع بخش",
                "زیربخش", "پاسخ سوال ۱", "پاسخ سوال ۲", "پاسخ سوال ۳",
                "پاسخ سوال ۴", "پاسخ سوال ۵", "تاریخ ثبت"]
    _style_header(ws4, headers4)

    form_type_fa = {
        "business": "کسب‌وکار",
        "social": "مسئولیت اجتماعی",
        "growth": "مسیر رشد",
    }

    row_idx = 2
    for user_id, records in section_forms.items():
        user_info = users.get(user_id, {})
        for record in records:
            ws4.cell(row=row_idx, column=1, value=row_idx - 1)
            ws4.cell(row=row_idx, column=2, value=user_info.get('telegram_id', user_id))
            ws4.cell(row=row_idx, column=3, value=user_info.get('first_name', ''))
            ws4.cell(row=row_idx, column=4, value=user_info.get('last_name', ''))
            ws4.cell(row=row_idx, column=5, value=form_type_fa.get(record.get('form_type', ''), record.get('form_type', '')))
            ws4.cell(row=row_idx, column=6, value=record.get('sub_title', ''))

            qa_list = record.get('qa', [])
            for i in range(5):
                value = qa_list[i]['answer'] if i < len(qa_list) else ''
                ws4.cell(row=row_idx, column=7 + i, value=value)

            ws4.cell(row=row_idx, column=12, value=record.get('submitted_at', ''))
            row_idx += 1

    # ===== شیت ۵: سفارش محصولات (تب «محصولات و خدمات سیناپس») =====
    ws5 = wb.create_sheet("سفارش محصولات")

    headers5 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "شماره تماس",
                "متن سفارش", "وضعیت", "تاریخ ثبت"]
    _style_header(ws5, headers5)

    row_idx = 2
    for user_id, orders in product_orders.items():
        user_info = users.get(user_id, {})
        for order in orders:
            ws5.cell(row=row_idx, column=1, value=row_idx - 1)
            ws5.cell(row=row_idx, column=2, value=user_info.get('telegram_id', user_id))
            ws5.cell(row=row_idx, column=3, value=user_info.get('first_name', ''))
            ws5.cell(row=row_idx, column=4, value=user_info.get('last_name', ''))
            ws5.cell(row=row_idx, column=5, value=user_info.get('phone', ''))
            ws5.cell(row=row_idx, column=6, value=order.get('order_text', ''))
            ws5.cell(row=row_idx, column=7, value=order.get('status', ''))
            ws5.cell(row=row_idx, column=8, value=order.get('submitted_at', ''))
            row_idx += 1

    # ===== شیت ۶: وضعیت و تاریخچه‌ی اشتراک =====
    ws6 = wb.create_sheet("وضعیت اشتراک")

    headers6 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "وضعیت فعلی",
                "سطح", "تاریخ درخواست", "تاریخ تایید", "تاریخ انقضا"]
    _style_header(ws6, headers6)

    for row, (user_id, sub) in enumerate(subscriptions.items(), 2):
        user_info = users.get(user_id, {})
        ws6.cell(row=row, column=1, value=row - 1)
        ws6.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws6.cell(row=row, column=3, value=user_info.get('first_name', ''))
        ws6.cell(row=row, column=4, value=user_info.get('last_name', ''))
        ws6.cell(row=row, column=5, value=status_labels_fa.get(sub.get('status', 'none'), sub.get('status', '')))
        ws6.cell(row=row, column=6, value=tier_labels.get(sub.get('tier'), '-'))
        ws6.cell(row=row, column=7, value=sub.get('requested_at', '-'))
        ws6.cell(row=row, column=8, value=sub.get('approved_at', '-'))
        ws6.cell(row=row, column=9, value=sub.get('expires_at', '-'))

    # ===== شیت ۷: خلاصه آمار =====
    ws7 = wb.create_sheet("خلاصه آمار")

    total_section_forms = sum(len(records) for records in section_forms.values())
    total_product_orders = sum(len(orders) for orders in product_orders.values())
    total_active_subs = sum(1 for s in subscriptions.values() if s.get("status") == "active")

    stats_data = [
        ["آمار کلی", ""],
        ["تعداد کل کاربران", len(users)],
        ["تعداد فرم‌های ارزیابی تکمیل شده", len(assessments)],
        ["تعداد پرسشنامه‌های تکمیل شده", sum(1 for u in surveys if len(surveys[u]) > 2)],
        ["تعداد فرم‌های بخش‌ها (کسب‌وکار/مسئولیت/رشد) تکمیل‌شده", total_section_forms],
        ["تعداد سفارش‌های ثبت‌شده از محصولات/خدمات", total_product_orders],
        ["تعداد اشتراک‌های فعال فعلی", total_active_subs],
        ["تاریخ تولید گزارش", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["آخرین ۵ کاربر", ""],
    ]

    for i, (user_id, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        stats_data.append([f"{i}. {name}", info.get('business_name', '')])

    for row, (key, value) in enumerate(stats_data, 1):
        ws7.cell(row=row, column=1, value=key)
        ws7.cell(row=row, column=2, value=value)

    ws7.column_dimensions['A'].width = 35
    ws7.column_dimensions['B'].width = 30

    wb.save(EXCEL_FILE)
    return EXCEL_FILE

# ==================== دستور /getexcel - دریافت فایل اکسل ====================
async def get_excel(update: Update, context):
    """ادمین می‌تواند با این دستور فایل اکسل کامل کاربران را دریافت کند"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    await update.message.reply_text("📊 در حال تولید فایل اکسل... لطفاً صبر کنید...")

    try:
        excel_file = generate_excel_report()
        with open(excel_file, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f'users_data_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                caption="📊 **گزارش کامل کاربران**\n\n"
                        f"📅 تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                        "📋 شامل: اطلاعات کاربران + فرم ارزیابی + پرسشنامه + فرم‌های بخش‌ها "
                        "+ سفارش محصولات + وضعیت اشتراک + آمار"
            )
    except Exception as e:
        await update.message.reply_text(f"⚠️ خطا در تولید فایل: {str(e)}")

# ==================== دستور /getdata - دریافت فایل‌های JSON خام ====================
async def get_data(update: Update, context):
    """ادمین می‌تواند فایل‌های JSON خام را دریافت کند"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    files_to_send = [
        (USERS_FILE, f'users_{datetime.now().strftime("%Y%m%d")}.json'),
        (SURVEY_FILE, f'survey_{datetime.now().strftime("%Y%m%d")}.json'),
        (ASSESSMENT_FILE, f'assessment_{datetime.now().strftime("%Y%m%d")}.json'),
        (SECTION_FORMS_FILE, f'section_forms_{datetime.now().strftime("%Y%m%d")}.json'),
        (SUBSCRIPTIONS_FILE, f'subscriptions_{datetime.now().strftime("%Y%m%d")}.json'),
        (PRODUCT_ORDERS_FILE, f'product_orders_{datetime.now().strftime("%Y%m%d")}.json'),
        (BANNED_USERS_FILE, f'banned_users_{datetime.now().strftime("%Y%m%d")}.json'),
    ]

    for file_path, filename in files_to_send:
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                await update.message.reply_document(f, filename=filename)

# ==================== توابع مشترک ساخت متن (برای دستور و دکمه‌ی پنل هر دو) ====================
def build_summary_text():
    """متن آمار کلی - هم در /summary و هم در دکمه‌ی پنل ادمین استفاده می‌شود"""
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    assessments = read_json(ASSESSMENT_FILE, {})
    section_forms = read_json(SECTION_FORMS_FILE, {})
    subscriptions = read_json(SUBSCRIPTIONS_FILE, {})
    product_orders = read_json(PRODUCT_ORDERS_FILE, {})

    completed_surveys = sum(1 for u in surveys if len(surveys[u]) > 2)
    total_section_forms = sum(len(records) for records in section_forms.values())
    total_orders = sum(len(orders) for orders in product_orders.values())
    active_subs = sum(1 for s in subscriptions.values() if s.get("status") == "active")
    today = datetime.now().strftime("%Y-%m-%d")
    today_users = sum(1 for u in users.values() if u.get("last_update", "").startswith(today))

    summary = f"📊 **آمار کلی:**\n\n"
    summary += f"👥 کل کاربران: {len(users)}\n"
    summary += f"📋 فرم‌های ارزیابی تکمیل شده: {len(assessments)}\n"
    summary += f"📊 پرسشنامه‌های تکمیل شده: {completed_surveys}\n"
    summary += f"🧩 فرم‌های بخش‌ها (کسب‌وکار/مسئولیت/رشد): {total_section_forms}\n"
    summary += f"🌱 سفارش‌های محصولات/خدمات: {total_orders}\n"
    summary += f"💎 اشتراک‌های فعال: {active_subs}\n"
    summary += f"🆕 کاربران امروز: {today_users}\n\n"
    summary += "**آخرین ۵ کاربر:**\n"

    for i, (uid, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        business = info.get('business_name', 'نامشخص')
        summary += f"{i}. {name} - {business}\n"

    return summary

def build_pending_text():
    """متن خلاصه‌ی موارد در انتظار بررسی - سفارش‌های محصول و درخواست‌های اشتراک"""
    from config import get_pending_items
    pending = get_pending_items()
    users = read_json(USERS_FILE, {})

    text = "⏳ **موارد در انتظار بررسی**\n\n"

    text += f"🌱 سفارش‌های محصول/خدمت ({len(pending['orders'])}):\n"
    if pending["orders"]:
        for uid, order in pending["orders"][:15]:
            info = users.get(uid, {})
            name = f"{info.get('first_name', '')} {info.get('last_name', '')}".strip() or "نامشخص"
            text += f"  • {name} (`{uid}`): {order.get('order_text', '')[:50]}\n"
    else:
        text += "  هیچ موردی وجود ندارد.\n"

    text += f"\n💳 درخواست‌های اشتراک ({len(pending['subscriptions'])}):\n"
    if pending["subscriptions"]:
        for uid, sub in pending["subscriptions"][:15]:
            info = users.get(uid, {})
            name = f"{info.get('first_name', '')} {info.get('last_name', '')}".strip() or "نامشخص"
            text += f"  • {name} (`{uid}`) - سطح: {sub.get('tier', '-')}\n"
    else:
        text += "  هیچ موردی وجود ندارد.\n"

    return text

def build_search_results_text(query_str):
    """متن نتایج جستجوی کاربر بر اساس نام/شماره/نام کسب‌وکار"""
    from config import search_users
    results = search_users(query_str)

    if not results:
        return f"🔍 برای «{query_str}» هیچ کاربری پیدا نشد."

    text = f"🔍 نتایج جستجو برای «{query_str}» ({len(results)} مورد):\n\n"
    for uid, info in results:
        text += (
            f"👤 {info.get('first_name', '')} {info.get('last_name', '')}\n"
            f"🆔 `{uid}`\n"
            f"📞 {info.get('phone', '-')}\n"
            f"🏢 {info.get('business_name', '-')}\n"
            f"➖➖➖➖➖\n"
        )
    return text

# ==================== دستور /summary - آمار خلاصه ====================
async def show_summary(update: Update, context):
    """نمایش آمار کلی کاربران برای ادمین"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    await update.message.reply_text(build_summary_text(), parse_mode='Markdown')

# ==================== دستور /pending - موارد در انتظار بررسی ====================
async def show_pending(update: Update, context):
    """نمایش لیست سفارش‌های محصول و درخواست‌های اشتراک در انتظار بررسی"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    await update.message.reply_text(build_pending_text(), parse_mode='Markdown')

# ==================== دستور /find - جستجوی کاربر ====================
async def find_user(update: Update, context):
    """جستجوی کاربر بر اساس نام، نام خانوادگی، نام کسب‌وکار یا شماره تماس"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    if not context.args:
        await update.message.reply_text("⚠️ فرمت صحیح:\n/find <نام یا شماره تماس>")
        return
    query_str = " ".join(context.args)
    await update.message.reply_text(build_search_results_text(query_str), parse_mode='Markdown')

# ==================== دستور /ban - مسدودسازی یک کاربر ====================
async def ban_user_cmd(update: Update, context):
    """مسدود کردن یک کاربر خاص (دیگر پاسخی از بات دریافت نمی‌کند)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    if not context.args:
        await update.message.reply_text("⚠️ فرمت صحیح:\n/ban <آیدی تلگرام کاربر>")
        return
    from config import ban_user
    target_id = context.args[0]
    ban_user(target_id)
    await update.message.reply_text(f"🚫 کاربر {target_id} مسدود شد و دیگر نمی‌تواند از بات استفاده کند.")

# ==================== دستور /unban - رفع مسدودیت یک کاربر ====================
async def unban_user_cmd(update: Update, context):
    """رفع مسدودیت یک کاربر خاص"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    if not context.args:
        await update.message.reply_text("⚠️ فرمت صحیح:\n/unban <آیدی تلگرام کاربر>")
        return
    from config import unban_user
    target_id = context.args[0]
    if unban_user(target_id):
        await update.message.reply_text(f"✅ مسدودیت کاربر {target_id} رفع شد.")
    else:
        await update.message.reply_text(f"ℹ️ کاربر {target_id} اصلاً مسدود نبود.")

# ==================== دستور /broadcast - ارسال پیام همگانی ====================
async def broadcast(update: Update, context):
    """ارسال پیام به تمام کاربران ثبت‌شده (فقط ادمین)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    users = read_json(USERS_FILE, {})
    if not users:
        await update.message.reply_text("📭 هیچ کاربری ثبت نشده!")
        return

    message_text = " ".join(context.args)
    if not message_text:
        await update.message.reply_text("⚠️ لطفاً پیام خود را وارد کنید.\nمثال: /broadcast سلام به همه!")
        return

    await update.message.reply_text(f"📤 ارسال پیام به {len(users)} کاربر...")

    success = 0
    for user_id in users.keys():
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"📢 **پیام از طرف مدیریت:**\n\n{message_text}"
            )
            success += 1
        except:
            pass

    await update.message.reply_text(f"✅ پیام به {success} کاربر ارسال شد.")

# ==================== دستور /reply - پاسخ مستقیم ادمین به یک کاربر خاص ====================
async def reply_to_user(update: Update, context):
    """
    ادمین می‌تواند با این دستور مستقیماً به یک کاربر خاص پیام بدهد - مثلاً
    برای پاسخ به سفارش محصول/خدمت، یا ارسال مبلغ و جزئیات پرداخت.
    فرمت دستور: /reply <آیدی تلگرام کاربر> <متن پیام>
    مثال:      /reply 123456789 سلام، هزینه‌ی این خدمت ۵۰۰ هزار تومان است.
    """
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    if len(context.args) < 2:
        await update.message.reply_text(
            "⚠️ فرمت صحیح دستور:\n"
            "/reply <آیدی تلگرام کاربر> <متن پیام>\n\n"
            "مثال:\n/reply 123456789 سلام، هزینه‌ی این خدمت ۵۰۰ هزار تومان است."
        )
        return

    target_id = context.args[0]
    message_text = " ".join(context.args[1:])

    try:
        await context.bot.send_message(
            chat_id=int(target_id),
            text=f"📩 پاسخ از سیناپس:\n\n{message_text}"
        )
        await update.message.reply_text(f"✅ پیام شما به کاربر {target_id} ارسال شد.")
        logger.info(f"✅ ادمین پیام مستقیم به کاربر {target_id} ارسال کرد")
    except ValueError:
        await update.message.reply_text("⚠️ آیدی تلگرام واردشده معتبر نیست (باید فقط عدد باشد).")
    except Exception as e:
        await update.message.reply_text(f"⚠️ خطا در ارسال پیام: {str(e)}")
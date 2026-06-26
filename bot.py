import os
import json
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters, CallbackQueryHandler
from telegram.error import TelegramError, BadRequest, Forbidden
from google import genai

# ==================== تنظیمات لاگینگ ====================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ==================== تنظیمات ====================
BOT_TOKEN = os.environ.get("BOT_TOKEN")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
ADMIN_ID = int(os.environ.get("ADMIN_ID", "8065571732"))
CHANNEL_ID = os.environ.get("CHANNEL_ID", "@synapse_os")

if not BOT_TOKEN:
    raise ValueError("❌ BOT_TOKEN در متغیرهای محیطی تنظیم نشده است!")

logger.info(f"✅ ربات با موفقیت راه‌اندازی شد!")
logger.info(f"📢 کانال: {CHANNEL_ID}")
logger.info(f"👑 ادمین: {ADMIN_ID}")

SYSTEM_PROMPT = (
    "تو یک مشاور کسب و کار حرفه‌ای هستی به نام مریم شهبازی. "
    "با لحنی گرم، دوستانه و حرفه‌ای پاسخ بده. "
    "همیشه به فارسی روان پاسخ بده. "
    "پاسخ‌هایت را کوتاه و مفید نگه دار."
)

try:
    if GEMINI_API_KEY:
        client = genai.Client(api_key=GEMINI_API_KEY)
        logger.info("✅ Gemini متصل شد.")
    else:
        client = None
        logger.warning("⚠️ GEMINI_API_KEY تنظیم نشده.")
except Exception as e:
    logger.error(f"خطا در راه‌اندازی Gemini: {e}")
    client = None

# فایل‌های JSON
USERS_FILE = "users.json"
SURVEY_FILE = "survey.json"
ASSESSMENT_FILE = "assessment.json"
EXCEL_FILE = "users_data.xlsx"

# ==================== اطلاعات پشتیبانی ====================
SUPPORT_INFO = {
    "phone": "09134525212",
    "email": "Shahbazimary1995@gmail.com",
    "telegram": "@malam_shahbazi",
    "hours": "تمام وقت",
    "response_time": "حداکثر تا 24 ساعت"
}

# ==================== توابع JSON ====================
def read_json(file_path, default={}):
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return default
    return default

def write_json(file_path, data):
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ==================== توابع ذخیره ====================
def save_user_info(user_id, info):
    users = read_json(USERS_FILE, {})
    if str(user_id) not in users:
        users[str(user_id)] = {}
    users[str(user_id)].update(info)
    users[str(user_id)]["telegram_id"] = str(user_id)
    users[str(user_id)]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json(USERS_FILE, users)
    logger.info(f"✅ اطلاعات کاربر {user_id} ذخیره شد")
    return True

def save_assessment(user_id, answers):
    assessments = read_json(ASSESSMENT_FILE, {})
    if str(user_id) not in assessments:
        assessments[str(user_id)] = {}
    assessments[str(user_id)].update(answers)
    assessments[str(user_id)]["submitted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json(ASSESSMENT_FILE, assessments)
    logger.info(f"✅ فرم ارزیابی کاربر {user_id} ذخیره شد")
    return True

def save_survey_answer(user_id, section, answer):
    surveys = read_json(SURVEY_FILE, {})
    if str(user_id) not in surveys:
        surveys[str(user_id)] = {}
    surveys[str(user_id)][section] = answer
    surveys[str(user_id)]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json(SURVEY_FILE, surveys)

def get_user_info(user_id):
    users = read_json(USERS_FILE, {})
    return users.get(str(user_id), {})

def is_user_registered(user_id):
    users = read_json(USERS_FILE, {})
    user_data = users.get(str(user_id), {})
    return user_data.get("first_name") is not None and user_data.get("first_name") != ""

def is_business_registered(user_id):
    users = read_json(USERS_FILE, {})
    user_data = users.get(str(user_id), {})
    return user_data.get("business_name") is not None and user_data.get("business_name") != ""

def has_completed_assessment(user_id):
    assessments = read_json(ASSESSMENT_FILE, {})
    return str(user_id) in assessments and len(assessments[str(user_id)]) > 5

# ==================== توابع بررسی عضویت ====================
async def is_member_of_channel(user_id, context):
    """بررسی عضویت کاربر در کانال با روش مطمئن"""
    try:
        chat_member = await context.bot.get_chat_member(
            chat_id=CHANNEL_ID,
            user_id=user_id
        )
        valid_statuses = ['member', 'administrator', 'creator']
        is_member = chat_member.status in valid_statuses
        logger.info(f"{'✅' if is_member else '❌'} کاربر {user_id} - وضعیت کانال: {chat_member.status}")
        return is_member

    except Forbidden as e:
        # بات ادمین کانال نیست یا دسترسی ندارد
        logger.warning(f"⛔ بات به کانال دسترسی ندارد (کاربر {user_id}): {e}")
        # اگر بات ادمین کانال نباشد، فرض می‌کنیم کاربر عضو است تا ربات کار کند
        return True

    except BadRequest as e:
        error_msg = str(e).lower()
        logger.error(f"❌ درخواست نامعتبر (کاربر {user_id}): {e}")
        if "user not found" in error_msg:
            return False
        if "chat not found" in error_msg:
            logger.error(f"⚠️ کانال {CHANNEL_ID} پیدا نشد! آیدی کانال را بررسی کنید.")
            return True  # اگر کانال پیدا نشد، بات را بلاک نکنیم
        return False

    except TelegramError as e:
        logger.error(f"⚠️ خطای تلگرام (کاربر {user_id}): {e}")
        return False

    except Exception as e:
        logger.error(f"⚠️ خطای ناشناخته در بررسی عضویت (کاربر {user_id}): {e}")
        return False


async def send_join_message(update: Update, context=None):
    """ارسال پیام الزام به عضویت با دکمه بررسی"""
    channel_url = f"https://t.me/{CHANNEL_ID.lstrip('@')}"
    join_button = InlineKeyboardMarkup([
        [InlineKeyboardButton("📢 عضویت در کانال", url=channel_url)],
        [InlineKeyboardButton("✅ بررسی عضویت", callback_data="check_membership")]
    ])

    message = (
        "⚠️ برای استفاده از ربات، ابتدا باید در کانال ما عضو شوید!\n\n"
        "📢 لطفاً روی دکمه زیر کلیک کرده و در کانال عضو شوید.\n\n"
        f"🆔 آیدی کانال: {CHANNEL_ID}\n\n"
        "✅ بعد از عضویت، روی دکمه «بررسی عضویت» کلیک کنید.\n\n"
        "💡 نکته: اگر قبلاً عضو شده‌اید، روی دکمه بررسی عضویت کلیک کنید تا دوباره چک شود."
    )

    if update.message:
        await update.message.reply_text(
            message,
            reply_markup=join_button,
            disable_web_page_preview=True
        )
    elif update.callback_query:
        await update.callback_query.message.reply_text(
            message,
            reply_markup=join_button,
            disable_web_page_preview=True
        )

# ==================== خروجی اکسل ====================
def generate_excel_report():
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    assessments = read_json(ASSESSMENT_FILE, {})
    
    wb = openpyxl.Workbook()
    
    # ===== شیت 1: اطلاعات کاربران =====
    ws1 = wb.active
    ws1.title = "اطلاعات کاربران"
    
    headers = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "تاریخ تولد", 
               "شماره تماس", "شهر", "نام کسب و کار", "آدرس", "راه معرفی", "تاریخ ثبت"]
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row, (user_id, info) in enumerate(users.items(), 2):
        ws1.cell(row=row, column=1, value=row-1)
        ws1.cell(row=row, column=2, value=info.get('telegram_id', user_id))
        ws1.cell(row=row, column=3, value=info.get('first_name', ''))
        ws1.cell(row=row, column=4, value=info.get('last_name', ''))
        ws1.cell(row=row, column=5, value=info.get('birth_date', ''))
        ws1.cell(row=row, column=6, value=info.get('phone', ''))
        ws1.cell(row=row, column=7, value=info.get('city', ''))
        ws1.cell(row=row, column=8, value=info.get('business_name', ''))
        ws1.cell(row=row, column=9, value=info.get('address', ''))
        ws1.cell(row=row, column=10, value=info.get('referral_source', ''))
        ws1.cell(row=row, column=11, value=info.get('last_update', ''))
    
    for col in range(1, len(headers)+1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # ===== شیت 2: فرم ارزیابی =====
    ws2 = wb.create_sheet("فرم_ارزیابی")
    
    assessment_headers = ["ردیف", "آیدی تلگرام", "نام و نام خانوادگی", "شماره تماس", "سن", 
                         "استان و شهر", "نقش فعلی", "حوزه تخصصی", "سه توانمندی", 
                         "دغدغه شغلی", "درآمد ماهیانه", "هدف یک ساله", "موانع اصلی", 
                         "مسئله حل شدنی", "نیاز اصلی", "نکته تکمیلی", "تاریخ ثبت"]
    
    for col, header in enumerate(assessment_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row, (user_id, answers) in enumerate(assessments.items(), 2):
        user_info = users.get(user_id, {})
        ws2.cell(row=row, column=1, value=row-1)
        ws2.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws2.cell(row=row, column=3, value=answers.get('full_name', ''))
        ws2.cell(row=row, column=4, value=answers.get('phone', ''))
        ws2.cell(row=row, column=5, value=answers.get('age', ''))
        ws2.cell(row=row, column=6, value=answers.get('location', ''))
        ws2.cell(row=row, column=7, value=answers.get('role', ''))
        ws2.cell(row=row, column=8, value=answers.get('field', ''))
        ws2.cell(row=row, column=9, value=answers.get('strengths', ''))
        ws2.cell(row=row, column=10, value=answers.get('challenge', ''))
        ws2.cell(row=row, column=11, value=answers.get('income', ''))
        ws2.cell(row=row, column=12, value=answers.get('goal', ''))
        ws2.cell(row=row, column=13, value=answers.get('obstacles', ''))
        ws2.cell(row=row, column=14, value=answers.get('solve_problem', ''))
        ws2.cell(row=row, column=15, value=answers.get('need', ''))
        ws2.cell(row=row, column=16, value=answers.get('note', ''))
        ws2.cell(row=row, column=17, value=answers.get('submitted_at', ''))
    
    for col in range(1, len(assessment_headers)+1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # ===== شیت 3: پرسشنامه =====
    ws3 = wb.create_sheet("پرسشنامه")
    
    survey_headers = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "نام کسب و کار",
                     "درباره کسب و کار", "محصولات و مزیت", "زیرساخت مجازی", 
                     "تیم", "فروش ماهیانه", "چالش اصلی", "نیاز مشاوره"]
    
    for col, header in enumerate(survey_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row, (user_id, answers) in enumerate(surveys.items(), 2):
        user_info = users.get(user_id, {})
        ws3.cell(row=row, column=1, value=row-1)
        ws3.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws3.cell(row=row, column=3, value=user_info.get('first_name', ''))
        ws3.cell(row=row, column=4, value=user_info.get('last_name', ''))
        ws3.cell(row=row, column=5, value=user_info.get('business_name', ''))
        ws3.cell(row=row, column=6, value=answers.get('about_business', ''))
        ws3.cell(row=row, column=7, value=answers.get('products', ''))
        ws3.cell(row=row, column=8, value=answers.get('infrastructure', ''))
        ws3.cell(row=row, column=9, value=answers.get('team', ''))
        ws3.cell(row=row, column=10, value=answers.get('sales', ''))
        ws3.cell(row=row, column=11, value=answers.get('problem', ''))
        ws3.cell(row=row, column=12, value=answers.get('consulting', ''))
    
    for col in range(1, len(survey_headers)+1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # ===== شیت 4: خلاصه آمار =====
    ws4 = wb.create_sheet("خلاصه آمار")
    
    stats_data = [
        ["آمار کلی", ""],
        ["تعداد کل کاربران", len(users)],
        ["تعداد فرم‌های ارزیابی تکمیل شده", len(assessments)],
        ["تعداد پرسشنامه‌های تکمیل شده", sum(1 for u in surveys if len(surveys[u]) > 2)],
        ["تاریخ تولید گزارش", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["آخرین ۵ کاربر", ""],
    ]
    
    for i, (user_id, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        stats_data.append([f"{i}. {name}", info.get('business_name', '')])
    
    for row, (key, value) in enumerate(stats_data, 1):
        ws4.cell(row=row, column=1, value=key)
        ws4.cell(row=row, column=2, value=value)
    
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 30
    
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

# ==================== منوها ====================
main_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🟢 بازار کار"), KeyboardButton("🔵 کسب‌وکار")],
    [KeyboardButton("🟣 مسئولیت اجتماعی"), KeyboardButton("🟠 مسیر رشد")],
    [KeyboardButton("🔴 لیدی لجستیک"), KeyboardButton("🌱 محصولات سیناپس")],
    [KeyboardButton("📖 راهنمای انتخاب مسیر"), KeyboardButton("🆔 اطلاعات شخصی")],
    [KeyboardButton("🏢 اطلاعات کسب و کار"), KeyboardButton("📊 پرسشنامه تخصصی")],
    [KeyboardButton("💬 مشاوره هوشمند"), KeyboardButton("📞 ارتباط با پشتیبانی")],
    [KeyboardButton("💳 ارسال فیش پرداخت")]
], resize_keyboard=True)

market_menu = ReplyKeyboardMarkup([
    [KeyboardButton("👤 کارجو"), KeyboardButton("💼 فریلنسر")],
    [KeyboardButton("🏢 کارفرما")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

business_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🌟 برند شخصی"), KeyboardButton("🚀 برند محصولی")],
    [KeyboardButton("🏛️ برند سازمانی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

social_menu = ReplyKeyboardMarkup([
    [KeyboardButton("❤️ نیک‌اندیش داخل ایران"), KeyboardButton("🌍 نیک‌اندیش خارج ایران")],
    [KeyboardButton("🤝 پروژه اجتماعی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

growth_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🧠 توسعه فردی"), KeyboardButton("🎯 توسعه شغلی")],
    [KeyboardButton("📈 توسعه اثر اجتماعی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

logistics_menu = ReplyKeyboardMarkup([
    [KeyboardButton("💰 استعلام قیمت"), KeyboardButton("🌍 تأمین‌کننده خارجی")],
    [KeyboardButton("📦 حمل و اسناد"), KeyboardButton("📈 فروش و بازاریابی")],
    [KeyboardButton("🎓 آموزش واردات"), KeyboardButton("🧭 مشاوره تخصصی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

back_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

def get_confirm_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تایید و ثبت", callback_data="confirm")],
        [InlineKeyboardButton("✏️ ویرایش اطلاعات", callback_data="edit")],
        [InlineKeyboardButton("❌ انصراف", callback_data="cancel")]
    ])

# ==================== سوالات ====================
personal_info_questions = [
    ("first_name", "👤 نام خود را وارد کنید:"),
    ("last_name", "👨‍👩‍👧 نام خانوادگی خود را وارد کنید:"),
    ("birth_date", "📅 تاریخ تولد (مثال: 1370/05/15):"),
    ("phone", "📞 شماره تماس موبایل:"),
    ("city", "🏙️ شهر خود را وارد کنید:"),
]

business_info_questions = [
    ("business_name", "🏢 نام کسب و کار خود را وارد کنید:"),
    ("address", "📍 آدرس شعب / دفتر مرکزی:"),
    ("referral_source", "📢 از چه طریقی با ما آشنا شدید؟"),
]

survey_questions = [
    ("about_business", "1️⃣ درباره کسب و کار خود بنویسید:\n(فعالیت، سابقه، اهداف)"),
    ("products", "2️⃣ چه محصولات یا خدماتی دارید؟\nمزیت شما نسبت به رقبا چیست؟"),
    ("infrastructure", "3️⃣ چه زیرساخت‌های مجازی دارید؟\n(سایت، اینستاگرام، اپلیکیشن، ...)"),
    ("team", "4️⃣ تیم شما شامل چه نیروهایی است؟\n(تعداد و تخصص)"),
    ("sales", "5️⃣ بالاترین فروش ماهیانه شما چقدر بوده؟"),
    ("problem", "6️⃣ مهمترین چالش یا مشکل فعلی شما چیست؟"),
    ("consulting", "7️⃣ در چه زمینه‌ای نیاز به مشاوره دارید؟")
]

assessment_questions = [
    ("full_name", "👤 نام و نام خانوادگی:"),
    ("phone", "📲 شماره تماس:"),
    ("age", "🎂 سن:"),
    ("location", "📍 استان و شهر محل سکونت:"),
    ("role", "1️⃣ امروز بیشتر خودت را در کدام نقش می‌بینی؟\n\n👤 کارجو\n💼 فریلنسر\n🏢 کارفرما"),
    ("field", "2️⃣ در چه حوزه یا تخصصی فعالیت می‌کنی یا دوست داری فعالیت کنی؟"),
    ("strengths", "3️⃣ سه توانمندی یا مهارتی که فکر می‌کنی نقطه قوت تو هستند را بنویس."),
    ("challenge", "4️⃣ این روزها مهم‌ترین دغدغه یا چالش شغلی تو چیست؟"),
    ("income", "5️⃣ در حال حاضر حدود درآمد ماهانه تو چقدر است؟"),
    ("goal", "6️⃣ دوست داری یک سال آینده از نظر شغلی و درآمدی کجا باشی؟"),
    ("obstacles", "7️⃣ فکر می‌کنی مهم‌ترین مانعی که بین تو و هدفت قرار گرفته چیست؟"),
    ("solve_problem", "8️⃣ اگر قرار باشد فقط یک مسئله از زندگی کاری تو حل شود، دوست داری آن مسئله چه باشد؟"),
    ("need", "9️⃣ در یک جمله بگو امروز بیشتر از هر چیز به چه کمکی نیاز داری؟"),
    ("note", "🔟 آیا نکته‌ای هست که فکر می‌کنی برای شناخت بهتر تو باید بدانیم؟\n(اختیاری)")
]

# ==================== وضعیت کاربران ====================
user_states = {}

def get_user_state(user_id):
    return user_states.get(user_id, {"section": None, "step": 0, "temp": {}})

def set_user_state(user_id, section, step=0, temp=None):
    user_states[user_id] = {
        "section": section,
        "step": step,
        "temp": temp if temp else {}
    }

def clear_user_state(user_id):
    if user_id in user_states:
        del user_states[user_id]

# ==================== نوتیفیکیشن به ادمین ====================
async def notify_admin(context, user_id, info, section_type="personal"):
    try:
        if section_type == "personal":
            message = f"🆕 **کاربر جدید ثبت‌نام کرد!**\n\n"
            message += f"🆔 آیدی تلگرام: `{user_id}`\n"
            message += f"👤 نام: {info.get('first_name', '')} {info.get('last_name', '')}\n"
            message += f"🏙️ شهر: {info.get('city', '')}\n"
            message += f"📞 شماره: {info.get('phone', '')}"
        elif section_type == "assessment":
            message = f"📋 **فرم ارزیابی جدید!**\n\n"
            message += f"🆔 آیدی تلگرام: `{user_id}`\n"
            message += f"👤 نام: {info.get('full_name', '')}\n"
            message += f"📞 شماره: {info.get('phone', '')}\n"
            message += f"🎂 سن: {info.get('age', '')}\n"
            message += f"📍 مکان: {info.get('location', '')}\n"
            message += f"💼 نقش: {info.get('role', '')}\n"
            message += f"📊 حوزه: {info.get('field', '')}"
        else:
            message = f"🏢 **اطلاعات کسب و کار جدید!**\n\n"
            message += f"🆔 آیدی تلگرام: `{user_id}`\n"
            message += f"👤 کاربر: {info.get('first_name', '')} {info.get('last_name', '')}\n"
            message += f"🏢 نام کسب و کار: {info.get('business_name', '')}\n"
            message += f"📍 آدرس: {info.get('address', '')}"
        
        await context.bot.send_message(chat_id=ADMIN_ID, text=message, parse_mode='Markdown')
    except Exception as e:
        logger.error(f"خطا در ارسال نوتیفیکیشن: {e}")

# ==================== دستور start ====================
async def start(update: Update, context):
    user_id = update.effective_user.id
    logger.info(f"📩 دستور /start از کاربر {user_id}")

    try:
        # بررسی عضویت در کانال
        is_member = await is_member_of_channel(user_id, context)

        if not is_member:
            await send_join_message(update, context)
            return

        # اگر عضو هست، ادامه بده
        user_info = get_user_info(user_id)
        logger.info(f"✅ کاربر {user_id} وارد شد - عضو کانال")

        if is_user_registered(user_id):
            welcome_msg = (
                f"✨ خوش برگشتی {user_info.get('first_name', '')} عزیز!\n\n"
                "به سیناپس خوش اومدی. 🌱😍\n"
                "هر آدمی در یکی از این مسیرها به دنبال رشد و توسعه برای ساختن یک ورژن بهتر از خودشه. "
                "تو از کجا میخوای شروع کنی؟\n\n"
                "🟢 بازار کار\n"
                "🔵 کسب‌وکار\n"
                "🟣 مسئولیت اجتماعی\n"
                "🟠 مسیر رشد\n"
                "🔴 لیدی لجستیک\n"
                "🌱 محصولات سیناپس\n\n"
                "لطفاً مسیر موردنظرت را انتخاب کن. 👇"
            )
        else:
            welcome_msg = (
                "سلام سلام\n"
                "شهبازی هستم، مریم 😍🌱\n"
                "اینجا قراره هویت کسب و کار و برند خودتون رو بسازید و روز به روز فروش بیشتری رو تجربه کنین.\n"
                "با من همراه باش\n\n"
                "به سیناپس خوش اومدی. 🌱😍\n"
                "هر آدمی در یکی از این مسیرها به دنبال رشد و توسعه برای ساختن یک ورژن بهتر از خودشه. "
                "تو از کجا میخوای شروع کنی؟\n\n"
                "🟢 بازار کار\n"
                "🔵 کسب‌وکار\n"
                "🟣 مسئولیت اجتماعی\n"
                "🟠 مسیر رشد\n"
                "🔴 لیدی لجستیک\n"
                "🌱 محصولات سیناپس\n\n"
                "لطفاً مسیر موردنظرت را انتخاب کن. 👇"
            )

        try:
            with open('images/welcome.jpg', 'rb') as photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=welcome_msg,
                    reply_markup=main_menu
                )
        except Exception:
            await update.message.reply_text(welcome_msg, reply_markup=main_menu)

    except Exception as e:
        logger.error(f"❌ خطا در هندلر start برای کاربر {user_id}: {e}", exc_info=True)
        try:
            await update.message.reply_text(
                "⚠️ خطایی رخ داد. لطفاً دوباره /start را بزنید.",
                reply_markup=main_menu
            )
        except Exception:
            pass

# ==================== تابع نمایش خلاصه ====================
def get_info_summary(info, section_type):
    if section_type == "personal":
        return f"""📝 **خلاصه اطلاعات شخصی شما:**

👤 نام: {info.get('first_name', '❌')}
👨‍👩‍👧 نام خانوادگی: {info.get('last_name', '❌')}
📅 تاریخ تولد: {info.get('birth_date', '❌')}
📞 شماره تماس: {info.get('phone', '❌')}
🏙️ شهر: {info.get('city', '❌')}

آیا اطلاعات صحیح است؟"""
    
    elif section_type == "business":
        return f"""🏢 **خلاصه اطلاعات کسب و کار شما:**

نام کسب و کار: {info.get('business_name', '❌')}
📍 آدرس: {info.get('address', '❌')}
📢 راه معرفی: {info.get('referral_source', '❌')}

آیا اطلاعات صحیح است؟"""
    
    return ""

# ==================== پردازش منو ====================
async def handle_menu(update: Update, context):
    user_id = update.effective_user.id
    text = update.message.text.strip()

    # بررسی عضویت
    if not await is_member_of_channel(user_id, context):
        await send_join_message(update, context)
        return

    state = get_user_state(user_id)
    section = state.get("section")
    step = state.get("step", 0)
    temp = state.get("temp", {})

    # ===== بازگشت به منو =====
    if text == "🔙 بازگشت به منوی اصلی":
        clear_user_state(user_id)
        await update.message.reply_text("🔹 به منوی اصلی بازگشتید 👇", reply_markup=main_menu)
        return

    # ===== منوهای اصلی =====
    menus = {
        "🟢 بازار کار": market_menu,
        "🔵 کسب‌وکار": business_menu,
        "🟣 مسئولیت اجتماعی": social_menu,
        "🟠 مسیر رشد": growth_menu,
        "🔴 لیدی لجستیک": logistics_menu,
    }
    if text in menus:
        await update.message.reply_text(f"{text}\n\nلطفاً یکی از گزینه‌های زیر را انتخاب کنید:", reply_markup=menus[text])
        return

    # ===== لیدی لجستیک - زیرمجموعه‌ها =====
    logistics_forms = {
        "💰 استعلام قیمت": (
            "💰 استعلام قیمت\n\n"
            "🌱 برای بررسی هزینه و زمان تحویل، لطفاً اطلاعات زیر را ارسال کنید:\n\n"
            "1️⃣ نام کالا\n"
            "2️⃣ کشور مبدأ\n"
            "3️⃣ تعداد / وزن / حجم سفارش\n"
            "4️⃣ شهر مقصد\n"
            "5️⃣ HS Code (در صورت اطلاع)\n"
            "6️⃣ فاکتور خرید (Proforma Invoice) در صورت وجود\n"
            "7️⃣ توضیحات تکمیلی\n\n"
            "📎 اگر HS Code را نمی‌دانید، نام، تصویر یا کاتالوگ محصول را ارسال کنید.\n\n"
            "⏰ کارشناسان لیدی لجستیک پس از بررسی با شما تماس خواهند گرفت."
        ),
        "🌍 تأمین‌کننده خارجی": (
            "🌍 تأمین‌کننده خارجی\n\n"
            "🌱 برای پیدا کردن تأمین‌کننده مناسب، لطفاً اطلاعات زیر را ارسال کنید:\n\n"
            "1️⃣ نام محصول\n"
            "2️⃣ کشور ترجیحی (در صورت وجود)\n"
            "3️⃣ حجم یا تعداد موردنیاز\n"
            "4️⃣ هدف شما از خرید چیست؟\n(مصرف شخصی / فروش / تولید)\n"
            "5️⃣ توضیحات تکمیلی\n\n"
            "📎 در صورت وجود، تصویر یا نمونه محصول را ارسال کنید.\n\n"
            "⏰ پس از بررسی، تأمین‌کنندگان مناسب معرفی خواهند شد."
        ),
        "📦 حمل و اسناد": (
            "📦 حمل و اسناد بازرگانی\n\n"
            "🌱 برای بررسی شرایط حمل و امور اسنادی، لطفاً اطلاعات زیر را ارسال کنید:\n\n"
            "1️⃣ نام کالا\n"
            "2️⃣ کشور مبدأ\n"
            "3️⃣ شهر مقصد\n"
            "4️⃣ وضعیت فعلی بار\n(آماده حمل / در حال خرید / نیاز به مشاوره)\n"
            "5️⃣ HS Code (در صورت اطلاع)\n"
            "6️⃣ اسناد موجود\n(پروفرما، پکینگ لیست، بارنامه و …)\n"
            "7️⃣ توضیحات تکمیلی\n\n"
            "📎 ارسال اسناد موجود به بررسی سریع‌تر کمک می‌کند.\n\n"
            "⏰ کارشناسان لیدی لجستیک درخواست شما را بررسی خواهند کرد."
        ),
        "📈 فروش و بازاریابی": (
            "📈 فروش و بازاریابی\n\n"
            "🌱 برای معرفی نیروهای فروش و توسعه بازار، لطفاً اطلاعات زیر را ارسال کنید:\n\n"
            "1️⃣ نام شرکت یا برند\n"
            "2️⃣ حوزه فعالیت\n"
            "3️⃣ شهر فعالیت\n"
            "4️⃣ نوع نیروی موردنیاز\n"
            "5️⃣ شرح کوتاه نیاز شما\n"
            "6️⃣ شماره تماس\n\n"
            "⏰ پس از بررسی، پیشنهادهای مناسب ارائه خواهد شد."
        ),
        "🎓 آموزش واردات": (
            "🎓 آموزش واردات\n\n"
            "🌱 برای معرفی مناسب‌ترین مسیر آموزشی، لطفاً اطلاعات زیر را ارسال کنید:\n\n"
            "1️⃣ آیا تجربه واردات دارید؟\n(بله / خیر)\n"
            "2️⃣ هدف شما چیست؟\n(شروع واردات / توسعه کسب‌وکار / یادگیری تخصصی)\n"
            "3️⃣ محصول یا صنعت موردعلاقه شما چیست؟\n"
            "4️⃣ مهم‌ترین سوال یا چالش شما چیست؟\n\n"
            "⏰ پس از بررسی، دوره یا مسیر آموزشی مناسب معرفی خواهد شد."
        ),
        "🧭 مشاوره تخصصی": (
            "🧭 مشاوره تخصصی\n\n"
            "🌱 برای هماهنگی جلسه مشاوره، لطفاً اطلاعات زیر را ارسال کنید:\n\n"
            "1️⃣ نام و نام خانوادگی\n"
            "2️⃣ حوزه فعالیت\n"
            "3️⃣ موضوع مشاوره\n"
            "4️⃣ مهم‌ترین مسئله یا سوال شما\n"
            "5️⃣ تاکنون چه اقداماتی انجام داده‌اید؟\n"
            "6️⃣ دوست دارید به چه نتیجه‌ای برسید؟\n"
            "7️⃣ شماره تماس\n\n"
            "⏰ پس از بررسی، زمان جلسه با شما هماهنگ خواهد شد."
        ),
    }
    if text in logistics_forms:
        set_user_state(user_id, "logistics_waiting", 0, {"service": text})
        await update.message.reply_text(logistics_forms[text], reply_markup=back_menu)
        await update.message.reply_text(
            "📝 لطفاً اطلاعات خواسته شده را در یک پیام ارسال کنید:",
            reply_markup=back_menu
        )
        return

    # ===== انتظار پاسخ لجستیک =====
    if section == "logistics_waiting":
        service = temp.get("service", "درخواست")
        user_info = get_user_info(user_id)
        first_name = user_info.get("first_name", "کاربر")
        admin_msg = (
            f"📋 درخواست جدید - {service}\n"
            f"👤 {first_name} {user_info.get('last_name','')}\n"
            f"🆔 {user_id}\n"
            f"📱 {user_info.get('phone','')}\n\n"
            f"💬 اطلاعات:\n{text}"
        )
        try:
            await context.bot.send_message(ADMIN_ID, admin_msg)
        except Exception:
            pass
        clear_user_state(user_id)
        await update.message.reply_text(
            "🌱 درخواست شما با موفقیت ثبت شد.\n\n"
            "اطلاعات ارسال‌شده توسط کارشناسان لیدی لجستیک بررسی می‌شود و در اولین فرصت "
            "برای هماهنگی و ارائه راهکار با شما تماس خواهیم گرفت.\n\n"
            "⏰ زمان پاسخگویی: حداکثر ۲ ساعت کاری.",
            reply_markup=main_menu
        )
        return

    # ===== محصولات سیناپس =====
    if text == "🌱 محصولات سیناپس":
        await update.message.reply_text(
            "🌱 محصولات سیناپس\n\n"
            "1️⃣ آموزش لینکدین\n"
            "2️⃣ ساخت رزومه موفق\n"
            "3️⃣ ساخت ربات تلگرام\n"
            "4️⃣ آموزش واردات\n"
            "5️⃣ طراحی آگهی استخدام\n"
            "6️⃣ طراحی بنر تبلیغاتی برای فروش\n"
            "7️⃣ طراحی لوگو و هویت آنلاین\n"
            "8️⃣ آموزش استخدام موفق\n"
            "9️⃣ فایل تامین‌کنندگان چینی\n"
            "🔟 افزایش درآمد ۳۰٪ کسب و کار (تضمینی)\n"
            "1️⃣1️⃣ فروش در دیجیکالا، ترب، باسلام\n\n"
            "برای دریافت اطلاعات هر محصول، شماره آن را بنویسید یا با پشتیبانی تماس بگیرید. 👇",
            reply_markup=main_menu
        )
        return

    # ===== زیرمجموعه‌های کسب‌وکار =====
    BUSINESS_QUESTIONS = {
        "🌟 برند شخصی": {
            "intro": "🌟 برند شخصی\n🌱 برای اینکه مسیرت را بهتر ببینیم:",
            "qs": [
                "1️⃣ امروز با چه تخصص یا مهارتی شناخته می‌شوی؟",
                "2️⃣ این روزها مهم‌ترین چالش تو چیست؟",
                "3️⃣ دوست داری یک سال دیگر چه تغییری در جایگاه حرفه‌ای تو ایجاد شده باشد؟",
                "4️⃣ فکر می‌کنی چه چیزی بیشتر از همه جلوی رشدت را گرفته است؟",
                "5️⃣ اگر سیناپس فقط یک کمک به تو بکند، دوست داری آن کمک چه باشد؟",
            ]
        },
        "🚀 برند محصولی": {
            "intro": "🚀 برند محصولی\n🌱 چند سؤال کوتاه برای شناخت بهتر کسب‌وکارت:",
            "qs": [
                "1️⃣ محصول یا خدمت شما چیست؟",
                "2️⃣ امروز مهم‌ترین چالش رشد شما چیست؟",
                "3️⃣ اگر یک سال دیگر همه چیز خوب پیش برود، چه اتفاقی افتاده است؟",
                "4️⃣ فکر می‌کنید بزرگ‌ترین مانع رسیدن به آن نقطه چیست؟",
                "5️⃣ اگر سیناپس فقط یک کمک به شما بکند، دوست دارید آن کمک چه باشد؟",
            ]
        },
        "🏛️ برند سازمانی": {
            "intro": "🏛️ برند سازمانی\n🌱 برای اینکه تصویر روشن‌تری از کسب‌وکار شما داشته باشیم:",
            "qs": [
                "1️⃣ نام سازمان یا کسب‌وکار شما چیست؟",
                "2️⃣ امروز مهم‌ترین مسئله‌ای که درگیر آن هستید چیست؟",
                "3️⃣ اگر یک سال دیگر به نتیجه مطلوب برسید، سازمان شما چه وضعیتی خواهد داشت؟",
                "4️⃣ چه چیزی بیشتر از همه مانع رسیدن به آن نقطه شده است؟",
                "5️⃣ اگر سیناپس فقط یک کمک به سازمان شما بکند، دوست دارید آن کمک چه باشد؟",
            ]
        },
    }

    SOCIAL_QUESTIONS = {
        "❤️ نیک‌اندیش داخل ایران": {
            "intro": "❤️ نیک‌اندیش داخل ایران\n🌱 هر اثر ماندگاری از یک دغدغه واقعی شروع می‌شود.",
            "qs": [
                "1️⃣ چه موضوع یا مسئله‌ای برای شما اهمیت بیشتری دارد؟",
                "2️⃣ دوست دارید در چه زمینه‌ای اثرگذار باشید؟\n(آموزش، اشتغال، سلامت، محیط زیست، زنان، کودکان و…)",
                "3️⃣ ترجیح می‌دهید چگونه مشارکت کنید؟\n(حمایت مالی، تخصص، زمان، ارتباطات)",
                "4️⃣ دوست دارید گزارش اثرگذاری شما شامل چه چیزی باشد؟",
                "5️⃣ اگر سیناپس فقط یک کمک به شما بکند، دوست دارید آن کمک چه باشد؟",
            ]
        },
        "🌍 نیک‌اندیش خارج ایران": {
            "intro": "🌍 نیک‌اندیش خارج از ایران\n🌱 فاصله جغرافیایی مانع اثرگذاری نیست.",
            "qs": [
                "1️⃣ چه موضوع یا مسئله‌ای برای شما اهمیت بیشتری دارد؟",
                "2️⃣ دوست دارید در چه زمینه‌ای اثرگذار باشید؟\n(آموزش، اشتغال، سلامت، محیط زیست، زنان، کودکان و…)",
                "3️⃣ ترجیح می‌دهید چگونه مشارکت کنید؟\n(حمایت مالی، انتقال دانش، ارتباطات، فرصت‌ها)",
                "4️⃣ دوست دارید گزارش اثرگذاری شما شامل چه چیزی باشد؟",
                "5️⃣ اگر سیناپس فقط یک کمک به شما بکند، دوست دارید آن کمک چه باشد؟",
            ]
        },
        "🤝 پروژه اجتماعی": {
            "intro": "🤝 پروژه اجتماعی\n🌱 هر پروژه اجتماعی برای حل یک مسئله واقعی شکل می‌گیرد.",
            "qs": [
                "1️⃣ پروژه شما روی حل چه مسئله‌ای تمرکز دارد؟",
                "2️⃣ امروز مهم‌ترین نیاز پروژه چیست؟\n(منابع مالی، نیروی انسانی، تخصص، رسانه، شبکه ارتباطی)",
                "3️⃣ اگر یک سال دیگر موفق شوید، چه تغییری ایجاد شده است؟",
                "4️⃣ چگونه اثرگذاری پروژه خود را اندازه‌گیری می‌کنید؟",
                "5️⃣ اگر سیناپس فقط یک کمک به پروژه شما بکند، دوست دارید آن کمک چه باشد؟",
            ]
        },
    }

    GROWTH_QUESTIONS = {
        "🧠 توسعه فردی": {
            "intro": "🧠 توسعه فردی\n🌱 گاهی قبل از هر تغییری، فقط نیاز داریم خودمان را واضح‌تر ببینیم.",
            "qs": [
                "1️⃣ این روزها بیشتر ذهنت درگیر چیست؟",
                "2️⃣ کدام بخش زندگی‌ات بیشتر از همه به توجه نیاز دارد؟",
                "3️⃣ اگر یک سال دیگر به حال خوب‌تری برسی، چه چیزی در زندگی‌ات تغییر کرده است؟",
                "4️⃣ فکر می‌کنی مهم‌ترین مانع رشد تو چیست؟",
                "5️⃣ اگر سیناپس فقط یک کمک به تو بکند، دوست داری آن کمک چه باشد؟",
            ]
        },
        "🎯 توسعه شغلی": {
            "intro": "💼 توسعه شغلی\n🌱 رشد شغلی فقط پیدا کردن کار نیست؛ پیدا کردن مسیر درست است.",
            "qs": [
                "1️⃣ امروز در چه مرحله‌ای از مسیر کاری خودت هستی؟",
                "2️⃣ این روزها مهم‌ترین دغدغه شغلی تو چیست؟",
                "3️⃣ دوست داری یک سال دیگر در چه جایگاهی باشی؟",
                "4️⃣ فکر می‌کنی چه چیزی بیشتر از همه جلوی رشد تو را گرفته است؟",
                "5️⃣ اگر سیناپس فقط یک کمک به تو بکند، دوست داری آن کمک چه باشد؟",
            ]
        },
        "📈 توسعه اثر اجتماعی": {
            "intro": "🤝 توسعه اثر اجتماعی\n🌱 هر اثر بزرگی از یک دغدغه کوچک شروع می‌شود.",
            "qs": [
                "1️⃣ چه مسئله‌ای در جامعه برای تو اهمیت بیشتری دارد؟",
                "2️⃣ دوست داری چه اثری از خودت به جا بگذاری؟",
                "3️⃣ تا امروز برای این دغدغه چه کاری انجام داده‌ای؟",
                "4️⃣ چه چیزی مانع اثرگذاری بیشتر تو شده است؟",
                "5️⃣ اگر سیناپس فقط یک کمک به تو بکند، دوست داری آن کمک چه باشد؟",
            ]
        },
    }

    SOCIAL_END_MSG = (
        "🌱 ممنون که دغدغه‌ات را با ما به اشتراک گذاشتی.\n\n"
        "سیناپس تلاش می‌کند میان آدم‌هایی که می‌خواهند اثر مثبتی خلق کنند و پروژه‌هایی که "
        "به حمایت نیاز دارند، پلی مطمئن بسازد.♥️\n\n"
        "اطلاعات شما بررسی می‌شود تا بتوانیم مناسب‌ترین فرصت‌های مشارکت، همکاری یا حمایت "
        "را به شما معرفی کنیم.\n\n"
        "همچنین در پروژه‌های منتخب، گزارش شفاف اثرگذاری و اعتبار اجتماعی مشارکت شما ارائه خواهد شد.\n\n"
        "⏰ حداکثر تا ۲۴ ساعت آینده با شما در ارتباط خواهیم بود."
    )

    GROWTH_END_MSG = (
        "🌱 ممنون که با صداقت پاسخ دادی.\n\n"
        "گاهی فقط چند سؤال کافی است تا تصویر شفاف‌تری از خودمان و مسیری که در آن هستیم ببینیم.♥️\n\n"
        "پاسخ‌های تو در سیناپس بررسی می‌شود تا گزارشی از وضعیت فعلی، ظرفیت‌های رشد، "
        "گره‌های مسیر و گام‌های پیشنهادی آماده شود.\n\n"
        "✨ آنچه امروز از تو می‌بینیم\n"
        "✨ نقاط قوت و ظرفیت‌ها\n"
        "✨ گره‌ها و موانع مسیر\n"
        "✨ فرصت‌های رشد\n"
        "✨ پیشنهاد گام بعدی\n\n"
        "⏰ زمان تحویل: حداکثر ۲۴ ساعت\n"
        "💳 هزینه گزارش شناخت: ۲۵۰ هزار تومان\n\n"
        "پس از پرداخت، فرآیند بررسی آغاز خواهد شد."
    )

    BUSINESS_END_MSG = (
        "🌱 ممنون که با صداقت پاسخ دادی.\n\n"
        "پاسخ‌های تو در سیناپس بررسی می‌شود تا گزارشی از وضعیت فعلی، نقاط قوت "
        "و فرصت‌های رشد آماده شود.\n\n"
        "✨ آنچه امروز از تو می‌بینیم\n"
        "✨ نقاط قوت و ظرفیت‌ها\n"
        "✨ گره‌ها و موانع مسیر\n"
        "✨ فرصت‌های رشد\n"
        "✨ پیشنهاد گام بعدی\n\n"
        "⏰ زمان تحویل: حداکثر ۲۴ ساعت\n"
        "💳 هزینه گزارش شناخت: ۲۵۰ هزار تومان\n\n"
        "پس از پرداخت، فیش را از طریق «💳 ارسال فیش پرداخت» ارسال کنید."
    )

    # ===== شروع فرم کسب‌وکار =====
    if text in BUSINESS_QUESTIONS:
        data = BUSINESS_QUESTIONS[text]
        set_user_state(user_id, "section_form", 0, {
            "form_type": "business",
            "sub": text,
            "intro": data["intro"],
            "qs": data["qs"],
            "end_msg": BUSINESS_END_MSG,
        })
        await update.message.reply_text(f"{data['intro']}\n\n{data['qs'][0]}", reply_markup=back_menu)
        return

    # ===== شروع فرم مسئولیت اجتماعی =====
    if text in SOCIAL_QUESTIONS:
        data = SOCIAL_QUESTIONS[text]
        set_user_state(user_id, "section_form", 0, {
            "form_type": "social",
            "sub": text,
            "intro": data["intro"],
            "qs": data["qs"],
            "end_msg": SOCIAL_END_MSG,
        })
        await update.message.reply_text(f"{data['intro']}\n\n{data['qs'][0]}", reply_markup=back_menu)
        return

    # ===== شروع فرم مسیر رشد =====
    if text in GROWTH_QUESTIONS:
        data = GROWTH_QUESTIONS[text]
        set_user_state(user_id, "section_form", 0, {
            "form_type": "growth",
            "sub": text,
            "intro": data["intro"],
            "qs": data["qs"],
            "end_msg": GROWTH_END_MSG,
        })
        await update.message.reply_text(f"{data['intro']}\n\n{data['qs'][0]}", reply_markup=back_menu)
        return

    # ===== پردازش فرم‌های section_form =====
    if section == "section_form":
        qs = temp.get("qs", [])
        answers = temp.get("answers", [])
        answers.append(text)
        temp["answers"] = answers

        if step + 1 < len(qs):
            set_user_state(user_id, "section_form", step + 1, temp)
            await update.message.reply_text(qs[step + 1], reply_markup=back_menu)
        else:
            # ذخیره و ارسال به ادمین
            user_info = get_user_info(user_id)
            first_name = user_info.get("first_name", "کاربر")
            sub = temp.get("sub", "")
            end_msg = temp.get("end_msg", "")
            admin_msg = (
                f"📋 فرم جدید — {sub}\n"
                f"👤 {first_name} {user_info.get('last_name','')}\n"
                f"🆔 {user_id} | 📱 {user_info.get('phone','')}\n\n"
            )
            for i, (q, a) in enumerate(zip(qs, answers), 1):
                admin_msg += f"❓ {q[:60]}\n💬 {a}\n\n"
            try:
                await context.bot.send_message(ADMIN_ID, admin_msg[:4000])
            except Exception:
                pass
            clear_user_state(user_id)
            await update.message.reply_text(end_msg, reply_markup=main_menu)
        return
    if text == "📖 راهنمای انتخاب مسیر":
        await update.message.reply_text(
            "📖 راهنمای انتخاب مسیر\n\n"
            "🟢 بازار کار — دنبال شغل، پروژه یا جذب نیرو هستی\n"
            "🔵 کسب‌وکار — می‌خواهی کسب‌وکارت را رشد بدهی\n"
            "🟣 مسئولیت اجتماعی — اثر مثبت اجتماعی می‌خواهی\n"
            "🟠 مسیر رشد — به دنبال خودشناسی و مسیر زندگی هستی\n"
            "🔴 لیدی لجستیک — خدمات واردات و صادرات\n"
            "🌱 محصولات سیناپس — ابزارها و دوره‌های آموزشی\n\n"
            "مسیر موردنظرت را انتخاب کن 👇",
            reply_markup=main_menu
        )
        return

    # ===== اطلاعات شخصی =====
    if text == "🆔 اطلاعات شخصی":
        if is_user_registered(user_id):
            user_info = get_user_info(user_id)
            first_name = user_info.get("first_name", "")
            edit_kb = InlineKeyboardMarkup([[InlineKeyboardButton("✏️ ویرایش اطلاعات", callback_data="edit_personal")]])
            await update.message.reply_text(
                f"سلام {first_name} عزیز! 👋\n\n"
                f"✅ اطلاعات شما قبلاً ثبت شده:\n"
                f"👤 {first_name} {user_info.get('last_name','')}\n"
                f"🏙️ {user_info.get('city','')}\n"
                f"📞 {user_info.get('phone','')}\n\n"
                f"برای ویرایش روی دکمه کلیک کنید 👇",
                reply_markup=edit_kb
            )
            return
        clear_user_state(user_id)
        set_user_state(user_id, "personal", 0, {})
        await update.message.reply_text(f"📝 ثبت اطلاعات شخصی\n\n{personal_info_questions[0][1]}", reply_markup=back_menu)
        return

    # ===== اطلاعات کسب و کار =====
    if text == "🏢 اطلاعات کسب و کار":
        if not is_user_registered(user_id):
            await update.message.reply_text("⚠️ ابتدا اطلاعات شخصی را ثبت کنید.", reply_markup=main_menu)
            return
        clear_user_state(user_id)
        set_user_state(user_id, "business", 0, {})
        await update.message.reply_text(f"🏢 ثبت اطلاعات کسب و کار\n\n{business_info_questions[0][1]}", reply_markup=back_menu)
        return

    # ===== پرسشنامه =====
    if text == "📊 پرسشنامه تخصصی":
        if not is_user_registered(user_id):
            await update.message.reply_text("⚠️ ابتدا اطلاعات شخصی را ثبت کنید.", reply_markup=main_menu)
            return
        clear_user_state(user_id)
        set_user_state(user_id, "survey", 0, {})
        await update.message.reply_text(f"📋 پرسشنامه تخصصی\n\n{survey_questions[0][1]}", reply_markup=back_menu)
        return

    # ===== مشاوره هوشمند =====
    if text == "💬 مشاوره هوشمند":
        if not is_user_registered(user_id):
            await update.message.reply_text("⚠️ ابتدا اطلاعات شخصی را ثبت کنید.", reply_markup=main_menu)
            return
        await update.message.reply_text(
            "💬 مشاوره هوشمند\n\nهر سوالی درباره برندسازی، بازاریابی، فروش و... داری بپرس 👇",
            reply_markup=back_menu
        )
        return

    # ===== ارسال فیش =====
    if text == "💳 ارسال فیش پرداخت":
        clear_user_state(user_id)
        set_user_state(user_id, "waiting_receipt", 0, {})
        await update.message.reply_text(
            "💳 پنل ارسال فیش پرداخت\n\n"
            "📸 تصویر فیش واریزی خود را ارسال کنید.\n\n"
            "• تصویر باید واضح و خوانا باشد\n"
            "• مبلغ و تاریخ باید مشخص باشد\n"
            "• بعد از ارسال حداکثر ۲۴ ساعت بررسی می‌شود",
            reply_markup=back_menu
        )
        return

    # ===== پشتیبانی =====
    if text == "📞 ارتباط با پشتیبانی":
        await update.message.reply_text(
            f"📞 ارتباط با پشتیبانی\n\n"
            f"📱 شماره تماس: {SUPPORT_INFO['phone']}\n"
            f"🆔 آیدی تلگرام: {SUPPORT_INFO['telegram']}\n"
            f"⏰ ساعات پاسخگویی: {SUPPORT_INFO['hours']}",
            reply_markup=back_menu
        )
        return

    # ===== پردازش فرم‌های مرحله‌ای =====

    if section == "assessment":
        if step < len(assessment_questions):
            field_name, _ = assessment_questions[step]
            temp[field_name] = text
            if step + 1 < len(assessment_questions):
                set_user_state(user_id, "assessment", step + 1, temp)
                await update.message.reply_text(assessment_questions[step + 1][1], reply_markup=back_menu)
            else:
                save_assessment(user_id, temp)
                await notify_admin(context, user_id, temp, "assessment")
                clear_user_state(user_id)
                await update.message.reply_text(
                    "🌱 ممنون که با حوصله پاسخ دادی.\n\n"
                    "📊 گزارش شناخت سیناپس شامل:\n"
                    "✨ آنچه امروز از تو می‌بینیم\n"
                    "✨ نقاط قوت و ظرفیت‌ها\n"
                    "✨ گره‌ها و موانع مسیر\n"
                    "✨ فرصت‌های رشد\n"
                    "✨ پیشنهاد گام بعدی\n\n"
                    "⏰ زمان تحویل: حداکثر ۲۴ ساعت\n"
                    "💳 هزینه گزارش شناخت: ۲۵۰ هزار تومان\n\n"
                    "پس از پرداخت، فیش را از طریق «💳 ارسال فیش پرداخت» ارسال کنید.",
                    reply_markup=main_menu
                )
        return

    if section == "personal":
        if step < len(personal_info_questions):
            field_name, _ = personal_info_questions[step]
            temp[field_name] = text
            if step + 1 < len(personal_info_questions):
                set_user_state(user_id, "personal", step + 1, temp)
                await update.message.reply_text(personal_info_questions[step + 1][1], reply_markup=back_menu)
            else:
                summary = get_info_summary(temp, "personal")
                set_user_state(user_id, "personal_confirm", 0, temp)
                await update.message.reply_text(summary, reply_markup=get_confirm_keyboard(), parse_mode="Markdown")
        return

    if section == "business":
        if step < len(business_info_questions):
            field_name, _ = business_info_questions[step]
            temp[field_name] = text
            if step + 1 < len(business_info_questions):
                set_user_state(user_id, "business", step + 1, temp)
                await update.message.reply_text(business_info_questions[step + 1][1], reply_markup=back_menu)
            else:
                summary = get_info_summary(temp, "business")
                set_user_state(user_id, "business_confirm", 0, temp)
                await update.message.reply_text(summary, reply_markup=get_confirm_keyboard(), parse_mode="Markdown")
        return

    if section == "survey":
        if step < len(survey_questions):
            field_name, _ = survey_questions[step]
            temp[field_name] = text
            if step + 1 < len(survey_questions):
                set_user_state(user_id, "survey", step + 1, temp)
                await update.message.reply_text(survey_questions[step + 1][1], reply_markup=back_menu)
            else:
                for key, value in temp.items():
                    save_survey_answer(user_id, key, value)
                clear_user_state(user_id)
                await update.message.reply_text(
                    "🌹 ممنون! پرسشنامه با موفقیت ثبت شد.\n\n"
                    "✅ ظرف ۴۸ ساعت کارشناسان با شما تماس می‌گیرند.",
                    reply_markup=main_menu
                )
        return

    if section == "waiting_receipt":
        await update.message.reply_text("📸 لطفاً تصویر فیش را ارسال کنید.", reply_markup=back_menu)
        return

    # ===== مشاوره هوشمند Gemini =====
    if not is_user_registered(user_id):
        await update.message.reply_text("⚠️ ابتدا اطلاعات شخصی را ثبت کنید.", reply_markup=main_menu)
        return

    if client is None:
        await update.message.reply_text("⚠️ سرویس هوشمند موقتاً در دسترس نیست.", reply_markup=back_menu)
        return

    user_info = get_user_info(user_id)
    first_name = user_info.get("first_name", "کاربر")
    await update.message.reply_text("⏳ در حال پردازش...")
    try:
        prompt = (
            f"{SYSTEM_PROMPT}\n\n"
            f"اطلاعات کاربر:\n"
            f"نام: {first_name} {user_info.get('last_name','')}\n"
            f"کسب‌وکار: {user_info.get('business_name','ثبت نشده')}\n"
            f"شهر: {user_info.get('city','ثبت نشده')}\n\n"
            f"سوال کاربر: {text}"
        )
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt
        )
        await update.message.reply_text(f"{first_name} عزیز،\n\n{response.text}", reply_markup=back_menu)
    except Exception as e:
        logger.error(f"Gemini Error: {e}", exc_info=True)
        await update.message.reply_text(
            f"⚠️ {first_name} عزیز، سرویس هوشمند موقتاً با مشکل مواجه شد.\n"
            f"خطا: {str(e)[:100]}\n\n"
            "لطفاً دوباره امتحان کنید.",
            reply_markup=back_menu
        )

# ==================== دکمه‌های اینلاین ====================
async def handle_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    data = query.data
    
    # ===== بررسی عضویت =====
    if data == "check_membership":
        is_member = await is_member_of_channel(user_id, context)

        if is_member:
            # پیام قبلی را حذف کن
            try:
                await query.edit_message_reply_markup(reply_markup=None)
            except Exception:
                pass

            user_info = get_user_info(user_id)

            if is_user_registered(user_id):
                welcome_msg = (
                    f"✨ خوش برگشتی {user_info.get('first_name', '')} عزیز!\n\n"
                    "به سیناپس خوش اومدی. 🌱😍\n"
                    "هر آدمی در یکی از این مسیرها به دنبال رشد و توسعه برای ساختن یک ورژن بهتر از خودشه. "
                    "تو از کجا میخوای شروع کنی؟\n\n"
                    "🟢 بازار کار\n"
                    "🔵 کسب‌وکار\n"
                    "🟣 مسئولیت اجتماعی\n"
                    "🟠 مسیر رشد\n"
                    "🔴 لیدی لجستیک\n"
                    "🌱 محصولات سیناپس\n\n"
                    "لطفاً مسیر موردنظرت را انتخاب کن. 👇"
                )
            else:
                welcome_msg = (
                    "سلام سلام\n"
                    "شهبازی هستم، مریم 😍🌱\n"
                    "اینجا قراره هویت کسب و کار و برند خودتون رو بسازید و روز به روز فروش بیشتری رو تجربه کنین.\n"
                    "با من همراه باش\n\n"
                    "به سیناپس خوش اومدی. 🌱😍\n"
                    "هر آدمی در یکی از این مسیرها به دنبال رشد و توسعه برای ساختن یک ورژن بهتر از خودشه. "
                    "تو از کجا میخوای شروع کنی؟\n\n"
                    "🟢 بازار کار\n"
                    "🔵 کسب‌وکار\n"
                    "🟣 مسئولیت اجتماعی\n"
                    "🟠 مسیر رشد\n"
                    "🔴 لیدی لجستیک\n"
                    "🌱 محصولات سیناپس\n\n"
                    "لطفاً مسیر موردنظرت را انتخاب کن. 👇"
                )

            # ارسال عکس خوش‌آمدگویی یا پیام متنی
            try:
                with open('images/welcome.jpg', 'rb') as photo:
                    await query.message.reply_photo(
                        photo=photo,
                        caption=welcome_msg,
                        reply_markup=main_menu
                    )
            except Exception:
                await query.message.reply_text(
                    welcome_msg,
                    reply_markup=main_menu
                )

        else:
            channel_url = f"https://t.me/{CHANNEL_ID.lstrip('@')}"
            join_button = InlineKeyboardMarkup([
                [InlineKeyboardButton("📢 عضویت در کانال", url=channel_url)],
                [InlineKeyboardButton("🔄 بررسی مجدد", callback_data="check_membership")]
            ])

            try:
                await query.edit_message_text(
                    f"❌ هنوز عضو کانال نشده‌اید!\n\n"
                    f"📢 لطفاً ابتدا در کانال {CHANNEL_ID} عضو شوید.\n"
                    f"✅ سپس روی دکمه «بررسی مجدد» کلیک کنید.",
                    reply_markup=join_button
                )
            except Exception:
                await query.message.reply_text(
                    f"❌ هنوز عضو کانال نشده‌اید!\n\n"
                    f"📢 لطفاً ابتدا در کانال {CHANNEL_ID} عضو شوید.\n"
                    f"✅ سپس روی دکمه «بررسی مجدد» کلیک کنید.",
                    reply_markup=join_button
                )
        return
    
    # ===== ویرایش اطلاعات شخصی =====
    if data == "edit_personal":
        clear_user_state(user_id)
        set_user_state(user_id, "personal", 0, {})
        try:
            await query.edit_message_reply_markup(reply_markup=None)
        except Exception:
            pass
        await query.message.reply_text(
            f"✏️ ویرایش اطلاعات شخصی\n\n{personal_info_questions[0][1]}",
            reply_markup=back_menu
        )
        return

    # ===== بقیه کدهای قبلی =====
    state = get_user_state(user_id)
    temp = state.get("temp", {})
    section = state.get("section", "")
    
    logger.info(f"کاربر {user_id} روی دکمه {data} کلیک کرد - بخش: {section}")
    
    if data == "confirm":
        try:
            if section == "personal_confirm":
                save_user_info(user_id, temp)
                await notify_admin(context, user_id, temp, "personal")
                first_name = temp.get('first_name', '')
                await query.edit_message_reply_markup(reply_markup=None)
                await query.message.reply_text(
                    f"✅ {first_name} عزیز، ثبت‌نام شما با موفقیت انجام شد! 🎉\n\n"
                    f"👤 نام: {first_name} {temp.get('last_name', '')}\n"
                    f"🏙️ شهر: {temp.get('city', '')}\n"
                    f"📞 شماره تماس: {temp.get('phone', '')}\n\n"
                    f"به سیناپس خوش اومدی! 🌱😍\n"
                    f"از منوی زیر مسیر موردنظرت را انتخاب کن 👇",
                    reply_markup=main_menu
                )
                clear_user_state(user_id)
                return
            
            elif section == "business_confirm":
                user_info = get_user_info(user_id)
                temp.update(user_info)
                save_user_info(user_id, temp)
                await notify_admin(context, user_id, temp, "business")
                
                await query.edit_message_reply_markup(reply_markup=None)
                
                await query.message.reply_text(
                    f"✅ **اطلاعات کسب و کار شما ثبت شد!** 🏢\n\n"
                    f"🏢 نام کسب و کار: {temp.get('business_name', '')}\n"
                    f"📍 آدرس: {temp.get('address', '')}\n"
                    f"📢 راه معرفی: {temp.get('referral_source', '')}\n\n"
                    f"📌 حالا می‌توانید:\n"
                    f"• پرسشنامه تخصصی را پر کنید\n"
                    f"• از مشاوره هوشمند استفاده کنید\n\n"
                    f"به منوی اصلی بازگشتید 👇",
                    reply_markup=main_menu,
                    parse_mode='Markdown'
                )
                clear_user_state(user_id)
                return
            
            else:
                await query.edit_message_reply_markup(reply_markup=None)
                await query.message.reply_text(
                    "⚠️ خطا در ثبت اطلاعات. لطفاً دوباره تلاش کنید.",
                    reply_markup=main_menu
                )
                clear_user_state(user_id)
                return
            
        except Exception as e:
            logger.error(f"خطا در تایید: {e}")
            await query.edit_message_reply_markup(reply_markup=None)
            await query.message.reply_text(
                f"⚠️ خطا در ثبت اطلاعات: {str(e)}\n\nلطفاً دوباره تلاش کنید.",
                reply_markup=main_menu
            )
            clear_user_state(user_id)
    
    elif data == "edit":
        if section == "personal_confirm":
            new_section = "personal"
            questions = personal_info_questions
            title = "✏️ ویرایش اطلاعات شخصی"
        elif section == "business_confirm":
            new_section = "business"
            questions = business_info_questions
            title = "✏️ ویرایش اطلاعات کسب و کار"
        else:
            await query.edit_message_text(
                "⚠️ خطا در ویرایش. لطفاً دوباره تلاش کنید.",
                reply_markup=main_menu
            )
            clear_user_state(user_id)
            return
        
        await query.edit_message_reply_markup(reply_markup=None)
        set_user_state(user_id, new_section, 0, {})
        await query.message.reply_text(
            f"{title}\n\n{questions[0][1]}",
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
    
    elif data == "cancel":
        clear_user_state(user_id)
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text(
            "❌ **ثبت‌نامه لغو شد.**\n\n"
            "اطلاعات شما ذخیره نشد.\n"
            "در صورت تمایل می‌توانید دوباره ثبت‌نام کنید.\n\n"
            "به منوی اصلی بازگشتید 👇",
            reply_markup=main_menu,
            parse_mode='Markdown'
        )

# ==================== دستورات ادمین ====================
async def get_excel(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    await update.message.reply_text("📊 در حال تولید فایل اکسل... لطفاً صبر کنید...")
    
    try:
        excel_file = generate_excel_report()
        with open(excel_file, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f'users_data_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                caption="📊 **گزارش کامل کاربران**\n\n"
                       f"📅 تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                       "📋 شامل: اطلاعات کاربران + فرم ارزیابی + پرسشنامه + آمار"
            )
    except Exception as e:
        await update.message.reply_text(f"⚠️ خطا در تولید فایل: {str(e)}")

async def get_data(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'rb') as f:
            await update.message.reply_document(f, filename=f'users_{datetime.now().strftime("%Y%m%d")}.json')
    
    if os.path.exists(SURVEY_FILE):
        with open(SURVEY_FILE, 'rb') as f:
            await update.message.reply_document(f, filename=f'survey_{datetime.now().strftime("%Y%m%d")}.json')
    
    if os.path.exists(ASSESSMENT_FILE):
        with open(ASSESSMENT_FILE, 'rb') as f:
            await update.message.reply_document(f, filename=f'assessment_{datetime.now().strftime("%Y%m%d")}.json')

async def show_summary(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    assessments = read_json(ASSESSMENT_FILE, {})
    
    completed_surveys = sum(1 for u in surveys if len(surveys[u]) > 2)
    today = datetime.now().strftime("%Y-%m-%d")
    today_users = sum(1 for u in users.values() if u.get("last_update", "").startswith(today))
    
    summary = f"📊 **آمار کلی:**\n\n"
    summary += f"👥 کل کاربران: {len(users)}\n"
    summary += f"📋 فرم‌های ارزیابی تکمیل شده: {len(assessments)}\n"
    summary += f"📊 پرسشنامه‌های تکمیل شده: {completed_surveys}\n"
    summary += f"🆕 کاربران امروز: {today_users}\n\n"
    summary += "**آخرین ۵ کاربر:**\n"
    
    for i, (uid, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        business = info.get('business_name', 'نامشخص')
        summary += f"{i}. {name} - {business}\n"
    
    await update.message.reply_text(summary, parse_mode='Markdown')

async def broadcast(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    users = read_json(USERS_FILE, {})
    if not users:
        await update.message.reply_text("📭 هیچ کاربری ثبت نشده!")
        return
    
    message_text = " ".join(context.args)
    if not message_text:
        await update.message.reply_text("⚠️ لطفاً پیام خود را وارد کنید.\nمثال: /broadcast سلام به همه!")
        return
    
    await update.message.reply_text(f"📤 ارسال پیام به {len(users)} کاربر...")
    
    success = 0
    for user_id in users.keys():
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"📢 **پیام از طرف مدیریت:**\n\n{message_text}"
            )
            success += 1
        except:
            pass
    
    await update.message.reply_text(f"✅ پیام به {success} کاربر ارسال شد.")

async def handle_photo(update: Update, context):
    """دریافت و فوروارد فیش پرداخت"""
    user_id = update.effective_user.id
    state = get_user_state(user_id)

    if state["section"] == "waiting_receipt":
        user_info = get_user_info(user_id)
        first_name = user_info.get('first_name', 'کاربر')
        caption = (
            f"💳 فیش پرداخت جدید\n"
            f"👤 {first_name} {user_info.get('last_name','')}\n"
            f"🆔 {user_id}\n"
            f"📱 {user_info.get('phone','ثبت نشده')}\n"
            f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        try:
            photo = update.message.photo[-1]
            await context.bot.send_photo(
                chat_id=ADMIN_ID,
                photo=photo.file_id,
                caption=caption
            )
            clear_user_state(user_id)
            await update.message.reply_text(
                f"✅ {first_name} عزیز، فیش پرداخت شما دریافت شد.\n\n"
                "🔍 بررسی شروع می‌شود و گزارش شناخت حداکثر ظرف ۲۴ ساعت ارسال خواهد شد. 🌱",
                reply_markup=main_menu
            )
        except Exception as e:
            logger.error(f"خطا در ارسال فیش: {e}")
            await update.message.reply_text(
                "⚠️ خطا در دریافت تصویر. لطفاً دوباره ارسال کنید.",
                reply_markup=back_menu
            )
    else:
        await update.message.reply_text(
            "📸 برای ارسال فیش پرداخت از دکمه «💳 ارسال فیش پرداخت» در منوی اصلی استفاده کنید.",
            reply_markup=main_menu
        )

# ==================== اجرا ====================
import traceback

async def error_handler(update, context):
    """هندلر سراسری خطا"""
    logger.error(f"❌ خطای سراسری: {context.error}", exc_info=context.error)
    if update and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "⚠️ خطایی رخ داد. لطفاً دوباره تلاش کنید یا /start را بزنید."
            )
        except Exception:
            pass

app = ApplicationBuilder().token(BOT_TOKEN).build()

app.add_error_handler(error_handler)
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("getexcel", get_excel))
app.add_handler(CommandHandler("getdata", get_data))
app.add_handler(CommandHandler("summary", show_summary))
app.add_handler(CommandHandler("broadcast", broadcast))
app.add_handler(CallbackQueryHandler(handle_callback))
app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu))

print("🤖 بات سیناپس با فرم ارزیابی روشن شد...")
print(f"📢 کانال اجباری: {CHANNEL_ID}")
print("📁 اطلاعات در فایل‌های JSON ذخیره می‌شوند")
print(f"👑 ادمین: {ADMIN_ID}")
print("📊 دستور /getexcel برای دریافت فایل اکسل")
app.run_polling()
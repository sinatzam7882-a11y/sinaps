# ==================== فایل گزارش اکسل و دستورات ادمین ====================
# این فایل شامل: تولید فایل اکسل، دستورات ادمین (/getexcel, /getdata, /summary, /broadcast) است
#
# نکته مهم: یک شیت جدید به نام «فرم‌های بخش‌ها» اضافه شده که پاسخ‌های
# کاربران به فرم‌های برند شخصی/محصولی/سازمانی، مسئولیت اجتماعی و مسیر رشد
# را نشان می‌دهد (این پاسخ‌ها قبلاً فقط برای ادمین پیامک می‌شدند و در
# اکسل ذخیره نمی‌شدند - الان در فایل section_forms.json ذخیره و در
# اکسل هم نمایش داده می‌شوند).

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from telegram import Update
from config import (
    read_json, USERS_FILE, SURVEY_FILE, ASSESSMENT_FILE, SECTION_FORMS_FILE,
    EXCEL_FILE, ADMIN_ID, logger
)

# ==================== استایل مشترک هدر شیت‌ها ====================
def _style_header(ws, headers):
    """رنگ، فونت و عرض ستون‌های هدر یک شیت را تنظیم می‌کند"""
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

# ==================== تولید فایل اکسل ====================
def generate_excel_report():
    """ساخت فایل اکسل با اطلاعات کاربران، فرم‌های ارزیابی، پرسشنامه، فرم‌های بخش‌ها و آمار"""
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    assessments = read_json(ASSESSMENT_FILE, {})
    section_forms = read_json(SECTION_FORMS_FILE, {})

    wb = openpyxl.Workbook()

    # ===== شیت ۱: اطلاعات کاربران =====
    ws1 = wb.active
    ws1.title = "اطلاعات کاربران"

    headers1 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "تاریخ تولد",
                "شماره تماس", "شهر", "نام کسب و کار", "آدرس", "راه معرفی", "تاریخ ثبت"]
    _style_header(ws1, headers1)

    for row, (user_id, info) in enumerate(users.items(), 2):
        ws1.cell(row=row, column=1, value=row - 1)
        ws1.cell(row=row, column=2, value=info.get('telegram_id', user_id))
        ws1.cell(row=row, column=3, value=info.get('first_name', ''))
        ws1.cell(row=row, column=4, value=info.get('last_name', ''))
        ws1.cell(row=row, column=5, value=info.get('birth_date', ''))
        ws1.cell(row=row, column=6, value=info.get('phone', ''))
        ws1.cell(row=row, column=7, value=info.get('city', ''))
        ws1.cell(row=row, column=8, value=info.get('business_name', ''))
        ws1.cell(row=row, column=9, value=info.get('address', ''))
        ws1.cell(row=row, column=10, value=info.get('referral_source', ''))
        ws1.cell(row=row, column=11, value=info.get('last_update', ''))

    # ===== شیت ۲: فرم ارزیابی بازار کار =====
    ws2 = wb.create_sheet("فرم_ارزیابی")

    headers2 = ["ردیف", "آیدی تلگرام", "نام و نام خانوادگی", "شماره تماس", "سن",
                "استان و شهر", "نقش فعلی", "حوزه تخصصی", "سه توانمندی",
                "دغدغه شغلی", "درآمد ماهیانه", "هدف یک ساله", "موانع اصلی",
                "مسئله حل شدنی", "نیاز اصلی", "نکته تکمیلی", "تاریخ ثبت"]
    _style_header(ws2, headers2)

    for row, (user_id, answers) in enumerate(assessments.items(), 2):
        user_info = users.get(user_id, {})
        ws2.cell(row=row, column=1, value=row - 1)
        ws2.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws2.cell(row=row, column=3, value=answers.get('full_name', ''))
        ws2.cell(row=row, column=4, value=answers.get('phone', ''))
        ws2.cell(row=row, column=5, value=answers.get('age', ''))
        ws2.cell(row=row, column=6, value=answers.get('location', ''))
        ws2.cell(row=row, column=7, value=answers.get('role', ''))
        ws2.cell(row=row, column=8, value=answers.get('field', ''))
        ws2.cell(row=row, column=9, value=answers.get('strengths', ''))
        ws2.cell(row=row, column=10, value=answers.get('challenge', ''))
        ws2.cell(row=row, column=11, value=answers.get('income', ''))
        ws2.cell(row=row, column=12, value=answers.get('goal', ''))
        ws2.cell(row=row, column=13, value=answers.get('obstacles', ''))
        ws2.cell(row=row, column=14, value=answers.get('solve_problem', ''))
        ws2.cell(row=row, column=15, value=answers.get('need', ''))
        ws2.cell(row=row, column=16, value=answers.get('note', ''))
        ws2.cell(row=row, column=17, value=answers.get('submitted_at', ''))

    # ===== شیت ۳: پرسشنامه تخصصی =====
    ws3 = wb.create_sheet("پرسشنامه")

    headers3 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "نام کسب و کار",
                "درباره کسب و کار", "محصولات و مزیت", "زیرساخت مجازی",
                "تیم", "فروش ماهیانه", "چالش اصلی", "نیاز مشاوره"]
    _style_header(ws3, headers3)

    for row, (user_id, answers) in enumerate(surveys.items(), 2):
        user_info = users.get(user_id, {})
        ws3.cell(row=row, column=1, value=row - 1)
        ws3.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws3.cell(row=row, column=3, value=user_info.get('first_name', ''))
        ws3.cell(row=row, column=4, value=user_info.get('last_name', ''))
        ws3.cell(row=row, column=5, value=user_info.get('business_name', ''))
        ws3.cell(row=row, column=6, value=answers.get('about_business', ''))
        ws3.cell(row=row, column=7, value=answers.get('products', ''))
        ws3.cell(row=row, column=8, value=answers.get('infrastructure', ''))
        ws3.cell(row=row, column=9, value=answers.get('team', ''))
        ws3.cell(row=row, column=10, value=answers.get('sales', ''))
        ws3.cell(row=row, column=11, value=answers.get('problem', ''))
        ws3.cell(row=row, column=12, value=answers.get('consulting', ''))

    # ===== شیت ۴: فرم‌های بخش‌ها (کسب‌وکار / مسئولیت اجتماعی / مسیر رشد) =====
    # هر ردیف یعنی یک فرم تکمیل‌شده توسط یک کاربر. چون هر کاربر می‌تواند
    # چند زیربخش مختلف را پر کند (مثلاً هم برند شخصی و هم توسعه شغلی)،
    # برای هرکدام یک ردیف جدا نوشته می‌شود. پاسخ هر سوال هم در یک ستون
    # مجزا (سوال ۱ تا ۵) قرار می‌گیرد تا خوانا باشد.
    ws4 = wb.create_sheet("فرم‌های بخش‌ها")

    headers4 = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "نوع بخش",
                "زیربخش", "پاسخ سوال ۱", "پاسخ سوال ۲", "پاسخ سوال ۳",
                "پاسخ سوال ۴", "پاسخ سوال ۵", "تاریخ ثبت"]
    _style_header(ws4, headers4)

    form_type_fa = {
        "business": "کسب‌وکار",
        "social": "مسئولیت اجتماعی",
        "growth": "مسیر رشد",
    }

    row_idx = 2
    for user_id, records in section_forms.items():
        user_info = users.get(user_id, {})
        for record in records:
            ws4.cell(row=row_idx, column=1, value=row_idx - 1)
            ws4.cell(row=row_idx, column=2, value=user_info.get('telegram_id', user_id))
            ws4.cell(row=row_idx, column=3, value=user_info.get('first_name', ''))
            ws4.cell(row=row_idx, column=4, value=user_info.get('last_name', ''))
            ws4.cell(row=row_idx, column=5, value=form_type_fa.get(record.get('form_type', ''), record.get('form_type', '')))
            ws4.cell(row=row_idx, column=6, value=record.get('sub_title', ''))

            qa_list = record.get('qa', [])
            for i in range(5):
                value = qa_list[i]['answer'] if i < len(qa_list) else ''
                ws4.cell(row=row_idx, column=7 + i, value=value)

            ws4.cell(row=row_idx, column=12, value=record.get('submitted_at', ''))
            row_idx += 1

    # ===== شیت ۵: خلاصه آمار =====
    ws5 = wb.create_sheet("خلاصه آمار")

    total_section_forms = sum(len(records) for records in section_forms.values())

    stats_data = [
        ["آمار کلی", ""],
        ["تعداد کل کاربران", len(users)],
        ["تعداد فرم‌های ارزیابی تکمیل شده", len(assessments)],
        ["تعداد پرسشنامه‌های تکمیل شده", sum(1 for u in surveys if len(surveys[u]) > 2)],
        ["تعداد فرم‌های بخش‌ها (کسب‌وکار/مسئولیت/رشد) تکمیل‌شده", total_section_forms],
        ["تاریخ تولید گزارش", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["آخرین ۵ کاربر", ""],
    ]

    for i, (user_id, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        stats_data.append([f"{i}. {name}", info.get('business_name', '')])

    for row, (key, value) in enumerate(stats_data, 1):
        ws5.cell(row=row, column=1, value=key)
        ws5.cell(row=row, column=2, value=value)

    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 30

    wb.save(EXCEL_FILE)
    return EXCEL_FILE

# ==================== دستور /getexcel - دریافت فایل اکسل ====================
async def get_excel(update: Update, context):
    """ادمین می‌تواند با این دستور فایل اکسل کامل کاربران را دریافت کند"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    await update.message.reply_text("📊 در حال تولید فایل اکسل... لطفاً صبر کنید...")

    try:
        excel_file = generate_excel_report()
        with open(excel_file, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f'users_data_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                caption="📊 **گزارش کامل کاربران**\n\n"
                        f"📅 تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                        "📋 شامل: اطلاعات کاربران + فرم ارزیابی + پرسشنامه + فرم‌های بخش‌ها + آمار"
            )
    except Exception as e:
        await update.message.reply_text(f"⚠️ خطا در تولید فایل: {str(e)}")

# ==================== دستور /getdata - دریافت فایل‌های JSON خام ====================
async def get_data(update: Update, context):
    """ادمین می‌تواند فایل‌های JSON خام را دریافت کند"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    files_to_send = [
        (USERS_FILE, f'users_{datetime.now().strftime("%Y%m%d")}.json'),
        (SURVEY_FILE, f'survey_{datetime.now().strftime("%Y%m%d")}.json'),
        (ASSESSMENT_FILE, f'assessment_{datetime.now().strftime("%Y%m%d")}.json'),
        (SECTION_FORMS_FILE, f'section_forms_{datetime.now().strftime("%Y%m%d")}.json'),
    ]

    for file_path, filename in files_to_send:
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                await update.message.reply_document(f, filename=filename)

# ==================== دستور /summary - آمار خلاصه ====================
async def show_summary(update: Update, context):
    """نمایش آمار کلی کاربران برای ادمین"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    assessments = read_json(ASSESSMENT_FILE, {})
    section_forms = read_json(SECTION_FORMS_FILE, {})

    completed_surveys = sum(1 for u in surveys if len(surveys[u]) > 2)
    total_section_forms = sum(len(records) for records in section_forms.values())
    today = datetime.now().strftime("%Y-%m-%d")
    today_users = sum(1 for u in users.values() if u.get("last_update", "").startswith(today))

    summary = f"📊 **آمار کلی:**\n\n"
    summary += f"👥 کل کاربران: {len(users)}\n"
    summary += f"📋 فرم‌های ارزیابی تکمیل شده: {len(assessments)}\n"
    summary += f"📊 پرسشنامه‌های تکمیل شده: {completed_surveys}\n"
    summary += f"🧩 فرم‌های بخش‌ها (کسب‌وکار/مسئولیت/رشد): {total_section_forms}\n"
    summary += f"🆕 کاربران امروز: {today_users}\n\n"
    summary += "**آخرین ۵ کاربر:**\n"

    for i, (uid, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        business = info.get('business_name', 'نامشخص')
        summary += f"{i}. {name} - {business}\n"

    await update.message.reply_text(summary, parse_mode='Markdown')

# ==================== دستور /broadcast - ارسال پیام همگانی ====================
async def broadcast(update: Update, context):
    """ارسال پیام به تمام کاربران ثبت‌شده (فقط ادمین)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return

    users = read_json(USERS_FILE, {})
    if not users:
        await update.message.reply_text("📭 هیچ کاربری ثبت نشده!")
        return

    message_text = " ".join(context.args)
    if not message_text:
        await update.message.reply_text("⚠️ لطفاً پیام خود را وارد کنید.\nمثال: /broadcast سلام به همه!")
        return

    await update.message.reply_text(f"📤 ارسال پیام به {len(users)} کاربر...")

    success = 0
    for user_id in users.keys():
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"📢 **پیام از طرف مدیریت:**\n\n{message_text}"
            )
            success += 1
        except:
            pass

    await update.message.reply_text(f"✅ پیام به {success} کاربر ارسال شد.")